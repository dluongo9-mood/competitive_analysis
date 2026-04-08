"""
THC/CBD Gummies Market Dashboard Builder

Loads Amazon, DTC, and Faire data, generates Plotly charts,
and outputs a single self-contained HTML dashboard.

Run:
    python3 build_dashboard.py
"""

import csv
import json
import re
import math
from pathlib import Path
from collections import Counter, defaultdict

import plotly.graph_objects as go
from plotly.subplots import make_subplots

# ── Data files ────────────────────────────────────────────────────────────────
AMAZON_CSV  = "amazon_gummies_plp.csv"
AMAZON_BRANDS_CSV = "amazon_brands.csv"
DTC_CSV     = "dtc_gummies.csv"
FAIRE_CSV   = "faire_gummies.csv"
TRAFFIC_CSV = "brand_traffic.csv"
WEEDMAPS_CSV = "weedmaps_gummies.csv"
BRIGHTFIELD_XLSX = "brightfield_hemp_thc.xlsx"

COLORS = {
    "Amazon":     "#FF9900",
    "DTC":        "#9B59B6",
    "Faire":      "#5B63FE",
    "Dispensary": "#2ECC71",
}

# ── Cannabinoid type classification ───────────────────────────────────────────
CANNABINOID_RULES = [
    ("Delta-9 THC",    [r"delta[\s-]*9", r"\bd9\b", r"\bΔ9"]),
    ("Delta-8 THC",    [r"delta[\s-]*8", r"\bd8\b", r"\bΔ8"]),
    ("CBD",            [r"\bcbd\b", r"cannabidiol"]),
    ("CBN",            [r"\bcbn\b", r"cannabinol"]),
    ("CBG",            [r"\bcbg\b", r"cannabigerol"]),
    ("THC-V",          [r"\bthcv\b", r"thc[\s-]*v\b"]),
    ("THC-P",          [r"\bthcp\b", r"thc[\s-]*p\b"]),
    ("HHC",            [r"\bhhc\b"]),
    ("THC-A",          [r"\bthca\b", r"thc[\s-]*a\b"]),
    ("Full Spectrum",  [r"full[\s-]*spectrum"]),
    ("Broad Spectrum", [r"broad[\s-]*spectrum"]),
]

# ── Form factor rules ─────────────────────────────────────────────────────────
FORM_FACTOR_RULES = [
    ("Gummy",      [r"gumm(?:y|ies)", r"chewable"]),
    ("Tincture",   [r"tincture", r"drops?", r"\boil\b"]),
    ("Vape",       [r"vape", r"cartridge", r"disposable", r"\bpen\b", r"\bcart\b"]),
    ("Edible",     [r"chocolate", r"cookie", r"brownie", r"candy", r"caramel", r"honey"]),
    ("Capsule",    [r"capsule", r"softgel", r"pill"]),
    ("Flower",     [r"flower", r"pre[\s-]*roll", r"joint", r"\bbud\b"]),
    ("Topical",    [r"cream", r"balm", r"salve", r"lotion", r"topical", r"roll[\s-]*on"]),
    ("Beverage",   [r"seltzer", r"drink", r"shot\b", r"sparkling", r"syrup"]),
]

# ── Use-case / effect classification ──────────────────────────────────────────
EFFECT_RULES = [
    ("Sleep",       [r"sleep", r"night", r"melatonin", r"zzz", r"rest\b", r"pm\b", r"bedtime"]),
    ("Relaxation",  [r"relax", r"calm", r"chill", r"stress", r"anxiety", r"unwind", r"ashwagandha"]),
    ("Pain",        [r"pain", r"relief", r"recover", r"inflammat", r"joint", r"muscle", r"aches?"]),
    ("Focus",       [r"focus", r"energy", r"uplift", r"clarity", r"boost", r"daytime", r"morning"]),
    ("Mood",        [r"mood", r"happy", r"euphori", r"bliss", r"vibe"]),
    ("Intimacy",    [r"intima", r"libido", r"arousal", r"sex"]),
    ("Cognitive",   [r"mushroom", r"lion.?s?\s*mane", r"nootropic", r"brain", r"mind\s*magic",
                     r"cordycep", r"reishi", r"chaga", r"cogniti"]),
    ("Immune",      [r"immune", r"turmeric", r"elderberry", r"vitamin\s*[cd]"]),
    ("Women's Health", [r"menopause", r"\bpms\b", r"hormone", r"period\s+(?:relief|support|pain)"]),
    ("Microdose",   [r"micro[\s-]*dos", r"low[\s-]*dos"]),
    ("Social",      [r"social", r"party", r"\bbuzz\b"]),
]

# ── Junk filter ───────────────────────────────────────────────────────────────
EXCLUDE_PATTERNS = [
    r"\bshirt\b", r"\bt-shirt\b", r"\bhoodie\b", r"\bhat\b", r"\bsocks?\b",
    r"\bsticker\b", r"\bposter\b", r"\bmug\b", r"\btote\b", r"\bpin\b",
    r"\bgift card\b", r"\bmerch\b", r"\bswag\b", r"\bgrinder\b",
    r"\brolling paper\b", r"\brolling tray\b", r"\bashtray\b",
    r"\bbattery\b(?!.*disposable)", r"\bcharger\b",
    r"\bdogs?\b", r"\bcats?\b", r"\bpets?\b",
    r"\bkratom\b", r"\binflatable\b", r"\bmystery box\b",
    r"\bnecklace\b", r"\bjewelry\b", r"\bearring\b", r"\bbracelet\b",
    r"\bcandle\b", r"\bsoap\b", r"\bbath bomb\b", r"\blip scrub\b",
    r"\bincense\b", r"\bpipe\b", r"\blighter\b",
    r"\bbook\b", r"\bcookbook\b", r"\bpuzzle\b", r"\bgame\b", r"\btoy\b",
    r"\bplush\b", r"\bscratcher\b",
    r"\bhemp protein\b", r"\bhemp hearts\b",
    r"\bbanner\b.*\bvinyl\b", r"\bvinyl\b.*\bbanner\b",
    r"\d+\"\s*[Xx×]\s*\d+\"\s*Banner",
]

# Faire brands that sell books/regular candy/non-THC — exclude all their products
FAIRE_EXCLUDE_BRANDS = {
    "simon & schuster", "microcosm publishing & distribution", "texas bookman",
    "chronicle books", "books by splitshops", "publishers group west",
    "grandpa joe's candy shop", "long island candy factory", "snacky candy",
    "rap snacks", "cow crack wholesale", "bebeto", "katjes",
    "beauty treats", "pet palette distribution",
    "everyday supply co", "hempz", "manitoba harvest", "bob's red mill",
    "nutiva", "navitas organics", "pacific natural foods",
    "nubian heritage", "dr. bronner's", "plant therapy", "moira cosmetics",
    "sassy cups", "good vibez collegiate", "camel threads", "lucky avocado",
    "natalie clare collections", "color shout", "plantlife", "organifi",
    "whynotnatural", "simply honest co.", "kurt s. adler, inc.", "kurt s. adler",
}

# Faire-specific: exclude hemp food/cosmetic products even if "hemp" matches
FAIRE_EXCLUDE_PRODUCTS = [
    r"moisturiz", r"body wash", r"body lotion", r"shampoo", r"conditioner",
    r"hemp seed", r"hemp oil(?!.*thc)(?!.*cbd)(?!.*delta)(?!.*gumm)",
    r"hemp hearts", r"hemp protein", r"hempseed", r"hemp powder",
    r"granola", r"overnight oats", r"cereal",
    r"hemp milk", r"non-dairy",
]
_faire_exclude_compiled = [re.compile(p, re.IGNORECASE) for p in FAIRE_EXCLUDE_PRODUCTS]
_exclude_compiled = [re.compile(p, re.IGNORECASE) for p in EXCLUDE_PATTERNS]

# Relevance gate: product must mention at least one of these keywords
RELEVANCE_KEYWORDS = [
    "thc", "cbd", "delta", "hemp", "gumm", "cbn", "cbg", "hhc",
    "cannabin", "edible", "tincture", "capsule", "vape", "topical",
    "seltzer", "full spectrum", "broad spectrum", "d8", "d9",
]


def is_excluded(text):
    return any(p.search(text) for p in _exclude_compiled)


def is_relevant(text):
    """Check if text contains at least one cannabinoid/product keyword."""
    t = (text or "").lower()
    return any(kw in t for kw in RELEVANCE_KEYWORDS)


# ── Amazon brand extraction ──────────────────────────────────────────────────
KNOWN_BRANDS = {
    "3chi": "3Chi", "delta extrax": "Delta Extrax", "hometown hero": "Hometown Hero",
    "koi cbd": "Koi CBD", "koi": "Koi CBD", "diamond cbd": "Diamond CBD",
    "exhale wellness": "Exhale Wellness", "exhale well": "Exhale Wellness",
    "budpop": "Budpop", "moonwlkr": "Moonwlkr", "binoid": "Binoid",
    "elyxr": "Elyxr", "urb": "Urb", "trehouse": "TRE House", "tre house": "TRE House",
    "trē house": "TRE House",
    "mr. hemp flower": "Mr. Hemp Flower", "viia hemp": "Viia Hemp",
    "delta munchies": "Delta Munchies", "otter space": "Otter Space",
    "sunday scaries": "Sunday Scaries", "cbdmd": "cbdMD",
    "charlotte's web": "Charlotte's Web", "charlottes web": "Charlotte's Web",
    "joy organics": "Joy Organics", "cornbread hemp": "Cornbread Hemp",
    "five cbd": "Five CBD", "5cbd": "Five CBD",
    "medterra": "Medterra", "batch cbd": "Batch CBD",
    "plain jane": "Plain Jane", "green roads": "Green Roads",
    "social cbd": "Social CBD", "tillmans tranquils": "Tillmans Tranquils",
    "cycling frog": "Cycling Frog", "wyld": "WYLD",
    "cheef botanicals": "Cheef Botanicals", "galaxy treats": "Galaxy Treats",
    "hemp bombs": "Hemp Bombs", "verma farms": "Verma Farms",
    "just cbd": "Just CBD", "feals": "Feals", "neurogan": "Neurogan",
    "secret nature": "Secret Nature", "area 52": "Area 52", "kanha": "Kanha",
    "hooloo": "HOOLOO", "afxmate": "AFXMATE", "olly": "OLLY",
    "r+r medicinals": "R+R Medicinals", "r + r medicinals": "R+R Medicinals",
    "soul cbd": "Soul CBD", "chill frog": "Chill Frog",
    "terra vita": "Terra Vita", "cbdfx": "CBDfx",
    "lazarus naturals": "Lazarus Naturals", "extract labs": "Extract Labs",
    "sunmed": "SunMed", "bloom hemp": "Bloom Hemp",
    "hempbombs": "Hemp Bombs", "hempgummies": "HempGummies",
    "slumber": "Slumber", "vida optima": "Vida Optima",
    "nature's script": "Nature's Script", "natures script": "Nature's Script",
    "sky wellness": "Sky Wellness", "mr hemp flower": "Mr. Hemp Flower",
    "edens herbals": "Eden's Herbals", "koi naturals": "Koi CBD",
    "penguin cbd": "Penguin CBD", "penguin": "Penguin CBD",
    "biobliss": "BioBliss", "nu-x": "Nu-X", "mystic labs": "Mystic Labs",
    "hometown hero cbd": "Hometown Hero",
    "wellution": "WELLUTION", "wellution hemp gummies": "WELLUTION",
    "wachray": "WACHRAY", "wachray nutrition": "WACHRAY",
    "hello mood": "Mood",
}

BAD_BRANDS = {
    "hemp", "delta", "potent", "extra", "premium", "organic", "natural", "mood",
    "gummies", "gummy", "new", "full", "broad", "spectrum", "the", "best",
    "advanced", "high", "strength", "super", "ultra", "max", "next",
    "pure", "original", "real", "total", "daily", "plant", "vitamatic",
    "hemp gummies", "hempgummies", "hemp gummy",
}


def extract_brand_from_title(title):
    """Try to extract brand name from an Amazon product title."""
    if not title:
        return None
    t_lower = title.lower().strip()

    # Check known brands (longest match first)
    for key, brand in sorted(KNOWN_BRANDS.items(), key=lambda x: -len(x[0])):
        if t_lower.startswith(key):
            return brand
        if key in t_lower:
            return brand

    # Try prefix extraction: first 1-3 capitalized words before a dash or common product word
    m = re.match(r'^([A-Z][A-Za-z0-9+\'\.]+(?:\s+[A-Z][A-Za-z0-9+\'\.]+){0,2})\s+[-–—|]', title)
    if m:
        candidate = m.group(1).strip()
        if candidate.lower() not in BAD_BRANDS and len(candidate) > 2:
            return candidate

    # Try: first word(s) before common product terms
    m2 = re.match(r'^([A-Z][A-Za-z0-9+\'\.]+(?:\s+[A-Z][A-Za-z0-9+\'\.]+){0,1})\s+(?:Hemp|Delta|CBD|THC|Gumm|Full|Broad|Organic|Premium)', title)
    if m2:
        candidate = m2.group(1).strip()
        if candidate.lower() not in BAD_BRANDS and len(candidate) > 2:
            return candidate

    return None


def extract_cannabinoids(text):
    found = []
    t = (text or "").lower()
    for name, patterns in CANNABINOID_RULES:
        if any(re.search(p, t) for p in patterns):
            found.append(name)
    return found


def infer_form_factor(text):
    if not text:
        return None
    t = text.lower()
    for label, patterns in FORM_FACTOR_RULES:
        for pat in patterns:
            if re.search(pat, t):
                return label
    return None


def extract_effects(text):
    found = []
    t = (text or "").lower()
    for name, patterns in EFFECT_RULES:
        if any(re.search(p, t) for p in patterns):
            found.append(name)
    return found


def parse_float(v):
    try:
        return float(v) if v not in (None, "", "None") else None
    except (ValueError, TypeError):
        return None


def parse_int(v):
    if v in (None, "", "None"):
        return None
    s = str(v).replace(",", "").strip()
    m = re.match(r'([\d.]+)\s*[Kk]\+?', s)
    if m:
        return int(float(m.group(1)) * 1000)
    m2 = re.match(r'(\d+)', s)
    return int(m2.group(1)) if m2 else None


def parse_sold(v):
    if not v:
        return None
    m = re.match(r'([\d.]+)\s*[Kk]\+?\s*bought', str(v))
    if m:
        return int(float(m.group(1)) * 1000)
    m2 = re.match(r'(\d+)\+?\s*bought', str(v))
    return int(m2.group(1)) if m2 else None


def read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def fig_to_html(fig):
    return fig.to_html(full_html=False, include_plotlyjs=False, config={"displayModeBar": True})


# ── Load data ─────────────────────────────────────────────────────────────────
def load_products():
    products = []
    excluded = 0

    # Load PDP brand data if available
    pdp_brands = {}
    if Path(AMAZON_BRANDS_CSV).exists():
        for r in read_csv(AMAZON_BRANDS_CSV):
            if r.get("brand"):
                pdp_brands[r["asin"]] = r["brand"]
        print(f"  PDP brands loaded: {len(pdp_brands)}")

    # Amazon — stricter: must mention a cannabinoid, not just "gummy"
    amz_irrelevant = 0
    AMZ_CANNABINOID_KW = ["thc", "cbd", "delta", "hemp", "cbn", "cbg", "hhc",
                          "cannabin", "full spectrum", "broad spectrum", "d8", "d9"]
    if Path(AMAZON_CSV).exists():
        for r in read_csv(AMAZON_CSV):
            title = r.get("title", "")
            if is_excluded(title):
                excluded += 1
                continue
            t_lower = title.lower()
            if not any(kw in t_lower for kw in AMZ_CANNABINOID_KW):
                amz_irrelevant += 1
                continue
            ff = infer_form_factor(title)
            cannabinoids = extract_cannabinoids(title)
            effects = extract_effects(title)
            asin = r.get("asin", "")
            raw_brand = pdp_brands.get(asin) or extract_brand_from_title(title)
            # Normalize through KNOWN_BRANDS
            brand = KNOWN_BRANDS.get(raw_brand.lower().strip(), raw_brand) if raw_brand else None
            products.append({
                "source": "Amazon", "id": asin,
                "brand": brand,
                "productName": title,
                "cannabinoids": cannabinoids,
                "effects": effects,
                "formFactor": ff or r.get("formFactor") or "Other",
                "price": parse_float(r.get("price")),
                "rating": parse_float(r.get("rating")),
                "reviewCount": parse_int(r.get("reviewCount")),
                "soldPastMonth": parse_sold(r.get("boughtPastMonth")),
                "url": r.get("url"),
            })
        amz_count = sum(1 for p in products if p['source'] == 'Amazon')
        amz_branded = sum(1 for p in products if p['source'] == 'Amazon' and p.get('brand'))
        print(f"  Amazon: {amz_count} products ({amz_branded} with brand, {amz_irrelevant} irrelevant filtered)")


    # DTC
    if Path(DTC_CSV).exists():
        dtc_excluded = 0
        dtc_irrelevant = 0
        for r in read_csv(DTC_CSV):
            name = r.get("productName", "")
            if is_excluded(name):
                dtc_excluded += 1
                continue
            price = parse_float(r.get("price"))
            if price is not None and price < 1.0:
                dtc_excluded += 1
                continue
            full_text = f"{name} {r.get('tags', '')} {r.get('productType', '')}"
            if not is_relevant(full_text):
                dtc_irrelevant += 1
                continue
            ff = infer_form_factor(full_text) or r.get("formFactor") or "Other"
            cannabinoids = extract_cannabinoids(full_text)
            if r.get("cannabinoids"):
                for c in r["cannabinoids"].split(","):
                    c = c.strip()
                    if c and c not in cannabinoids:
                        cannabinoids.append(c)
            effects = extract_effects(full_text)
            products.append({
                "source": "DTC", "id": r.get("productId"),
                "brand": r.get("brand"),
                "productName": name,
                "cannabinoids": cannabinoids,
                "effects": effects,
                "formFactor": ff,
                "price": price,
                "rating": None,
                "reviewCount": None,
                "soldPastMonth": None,
                "url": r.get("url"),
            })
        if dtc_excluded or dtc_irrelevant:
            print(f"  DTC: filtered {dtc_excluded} excluded + {dtc_irrelevant} irrelevant")
        print(f"  DTC: {sum(1 for p in products if p['source']=='DTC')} products")

    # Faire
    if Path(FAIRE_CSV).exists():
        faire_excluded = 0
        faire_irrelevant = 0
        for r in read_csv(FAIRE_CSV):
            name = r.get("name", "")
            brand = r.get("brand", "")
            if brand.lower() in FAIRE_EXCLUDE_BRANDS:
                faire_excluded += 1
                continue
            if is_excluded(name):
                faire_excluded += 1
                continue
            if any(p.search(name) for p in _faire_exclude_compiled):
                faire_excluded += 1
                continue
            if not is_relevant(name):
                faire_irrelevant += 1
                continue
            # Stricter Faire relevance: require specific cannabinoid keyword
            # "delta" alone matches Greek sororities, "hemp" alone matches food/cosmetics
            name_lower = name.lower()
            FAIRE_STRONG_KW = ["thc", "cbd", "cbn", "cbg", "hhc", "cannabin",
                               "delta 8", "delta 9", "delta-8", "delta-9", "d8", "d9",
                               "full spectrum", "broad spectrum", "thca", "thcp", "thcv"]
            has_strong_kw = any(k in name_lower for k in FAIRE_STRONG_KW)
            # If no strong keyword, only keep if both "hemp"/"delta" AND a product form word
            if not has_strong_kw:
                has_form = any(k in name_lower for k in ["gumm", "edible", "tincture",
                               "capsule", "vape", "topical", "seltzer", "softgel"])
                has_weak = any(k in name_lower for k in ["hemp", "delta"])
                if not (has_weak and has_form):
                    faire_irrelevant += 1
                    continue
            ff = infer_form_factor(name) or "Other"
            cannabinoids = extract_cannabinoids(name)
            effects = extract_effects(name)
            products.append({
                "source": "Faire", "id": r.get("productToken"),
                "brand": r.get("brand"),
                "productName": name,
                "cannabinoids": cannabinoids,
                "effects": effects,
                "formFactor": ff,
                "price": parse_float(r.get("retailPrice")),
                "rating": parse_float(r.get("rating")),
                "reviewCount": parse_int(r.get("reviewCount")),
                "soldPastMonth": None,
                "url": r.get("url"),
            })
        if faire_excluded or faire_irrelevant:
            print(f"  Faire: filtered {faire_excluded} excluded + {faire_irrelevant} irrelevant")
        print(f"  Faire: {sum(1 for p in products if p['source']=='Faire')} products")

    # Weedmaps (dispensary channel)
    if Path(WEEDMAPS_CSV).exists():
        wm_count = 0
        for r in read_csv(WEEDMAPS_CSV):
            name = r.get("productName", "")
            if not name or is_excluded(name):
                continue
            ff = infer_form_factor(name)
            if not ff:
                ff = "Gummy"  # Weedmaps data is pre-filtered to gummies
            cannabinoids = extract_cannabinoids(name)
            effects = extract_effects(name)
            price_str = r.get("price")
            price = parse_float(price_str) if price_str else None
            products.append({
                "source": "Dispensary", "id": r.get("url") or name,
                "brand": r.get("brand"),
                "productName": name,
                "cannabinoids": cannabinoids,
                "effects": effects,
                "formFactor": ff,
                "price": price,
                "rating": parse_float(r.get("rating")),
                "reviewCount": parse_float(r.get("reviewCount")),
                "soldPastMonth": None,
                "url": r.get("url"),
            })
            wm_count += 1
        if wm_count:
            print(f"  Weedmaps: {wm_count} products")

    if excluded:
        print(f"  Excluded {excluded} junk products total")

    # Deduplicate by (source, id)
    seen = set()
    deduped = []
    for p in products:
        key = (p["source"], p["id"])
        if key not in seen:
            seen.add(key)
            deduped.append(p)

    # Filter to gummies only — this is a gummies market dashboard
    before_ff = len(deduped)
    deduped = [p for p in deduped if p.get("formFactor") == "Gummy"]
    print(f"  {before_ff:,} → {len(deduped):,} after gummy-only filter")
    print(f"  {len(deduped):,} products loaded")
    return deduped


def load_traffic_data():
    """Load SimilarWeb traffic estimates for DTC brands."""
    traffic = {}
    if Path(TRAFFIC_CSV).exists():
        for r in read_csv(TRAFFIC_CSV):
            domain = r.get("domain", "")
            visits = parse_float(r.get("monthlyVisits"))
            if domain and visits:
                traffic[domain] = {
                    "brand": r.get("brand"),
                    "monthlyVisits": visits,
                    "bounceRate": parse_float(r.get("bounceRate")),
                    "globalRank": parse_float(r.get("globalRank")),
                }
    return traffic


def load_brightfield_data():
    """Load Brightfield Group hemp-derived THC market data from Excel."""
    if not Path(BRIGHTFIELD_XLSX).exists():
        return None
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not installed — skipping Brightfield data")
        return None

    wb = openpyxl.load_workbook(BRIGHTFIELD_XLSX, data_only=True)
    bf = {}

    # Sheet 1: Product Type Market Size (USD, 2020-2030)
    ws = wb["Product Type Market Size"]
    years = [int(c.value) for c in ws[5][1:12] if c.value]  # row 5 = header with years
    product_types = {}
    for row in ws.iter_rows(min_row=6, max_row=14, values_only=False):
        name = row[0].value
        if not name or not isinstance(name, str):
            continue
        vals = []
        for c in row[1:12]:
            v = c.value
            vals.append(float(v) if v is not None else None)
        product_types[name.strip()] = vals
    bf["product_type_years"] = years
    bf["product_type_sizes"] = product_types

    # Sheet 2: Cannabinoid Market Share (%, 2022-2025)
    ws2 = wb["Cannabinoids Market Share"]
    cb_years = [int(c.value) for c in ws2[5][1:5] if c.value]
    cannabinoids = {}
    for row in ws2.iter_rows(min_row=6, max_row=15, values_only=False):
        name = row[0].value
        if not name or not isinstance(name, str):
            continue
        vals = []
        for c in row[1:5]:
            v = c.value
            vals.append(float(v) * 100 if v is not None else None)  # convert to %
        cannabinoids[name.strip()] = vals
    bf["cannabinoid_years"] = cb_years
    bf["cannabinoid_shares"] = cannabinoids

    # Sheet 3: Company Market Share (%, 2022-2025)
    ws3 = wb["Company Market Share"]
    co_years = [int(c.value) for c in ws3[5][1:5] if c.value]
    companies = {}
    for row in ws3.iter_rows(min_row=6, max_row=43, values_only=False):
        name = row[0].value
        if not name or not isinstance(name, str):
            continue
        vals = []
        for c in row[1:5]:
            v = c.value
            vals.append(float(v) * 100 if v is not None else None)  # convert to %
        companies[name.strip()] = vals
    bf["company_years"] = co_years
    bf["company_shares"] = companies

    # Extract key numbers for context
    gummies_vals = product_types.get("Gummies", [])
    total_2025 = sum(v[5] for v in product_types.values() if v and len(v) > 5 and v[5] is not None)
    bf["gummies_2025"] = gummies_vals[5] if len(gummies_vals) > 5 else None
    bf["gummies_2024"] = gummies_vals[4] if len(gummies_vals) > 4 else None
    bf["gummies_2030"] = gummies_vals[10] if len(gummies_vals) > 10 else None
    bf["total_market_2025"] = total_2025

    print(f"  Brightfield: {len(product_types)} product types, {len(cannabinoids)} cannabinoids, {len(companies)} companies")
    if bf.get("gummies_2025"):
        print(f"  Gummies market: ${bf['gummies_2025']/1e9:.2f}B (2025), total market: ${total_2025/1e9:.2f}B")

    return bf


def compute_market_size(products, traffic):
    """Compute market size estimates across all channels."""
    # Amazon: measured revenue from soldPastMonth * price
    amz_measured = 0
    amz_measured_products = 0
    amz_total = 0
    for p in products:
        if p["source"] == "Amazon":
            amz_total += 1
            if p.get("soldPastMonth") and p.get("price"):
                amz_measured += p["price"] * p["soldPastMonth"]
                amz_measured_products += 1

    # Amazon: extrapolate for products without sold badges
    amz_no_badge = amz_total - amz_measured_products
    # Use actual median price of unbadged Amazon products for accuracy
    unbadged_prices = [p["price"] for p in products
                       if p["source"] == "Amazon" and not p.get("soldPastMonth") and p.get("price")]
    avg_amz_price = sorted(unbadged_prices)[len(unbadged_prices) // 2] if unbadged_prices else 29
    amz_extrapolated = amz_no_badge * 30 * avg_amz_price  # assume 30 units/mo avg for unbadged

    # DTC: estimate from traffic × per-brand conversion model
    # For each brand:
    #   1. Gummy share = % of brand's catalog that are gummies
    #   2. Median gummy price → AOV = median price × 1.3 (cart factor)
    #   3. Conversion rate = adjusted by bounce rate (low bounce → higher conv)
    #   4. Gummy revenue = visits × gummy_share × conv_rate × AOV
    dtc_revenue = {}
    dtc_brands_in_products = set()
    dtc_products_count = 0

    # Build per-brand product stats from RAW DTC CSV (not gummy-filtered products)
    # This gives us the true gummy share of each brand's catalog
    brand_stats = defaultdict(lambda: {"total": 0, "gummy": 0, "gummy_prices": []})
    if Path(DTC_CSV).exists():
        for r in read_csv(DTC_CSV):
            brand = r.get("brand", "").strip()
            if not brand:
                continue
            brand_stats[brand]["total"] += 1
            name = (r.get("productName") or "").lower()
            ff = (r.get("formFactor") or "").lower()
            tags = (r.get("tags") or "").lower()
            full = f"{name} {ff} {tags}"
            if "gumm" in full:
                brand_stats[brand]["gummy"] += 1
                price = parse_float(r.get("price"))
                if price and price > 1:
                    brand_stats[brand]["gummy_prices"].append(price)

    # Count DTC products (from gummy-filtered list)
    for p in products:
        if p["source"] == "DTC":
            dtc_products_count += 1
            if p.get("brand"):
                dtc_brands_in_products.add(p["brand"])

    # Generic domains that might get non-cannabis traffic — no discount needed for mood.com
    # (confirmed to be a cannabis-first brand despite generic domain)
    GENERIC_DOMAIN_DISCOUNT = {}

    for domain, data in traffic.items():
        brand = data.get("brand", "")
        visits = data.get("monthlyVisits", 0)
        bounce = data.get("bounceRate") or 0.5
        if isinstance(bounce, str):
            try:
                bounce = float(bounce)
            except (ValueError, TypeError):
                bounce = 0.5
        if not visits or not brand:
            continue

        bs = brand_stats.get(brand, {"total": 0, "gummy": 0, "gummy_prices": []})

        # Gummy catalog share (default 50% if no product data)
        gummy_catalog_pct = bs["gummy"] / max(bs["total"], 1) if bs["total"] > 0 else 0.5
        if gummy_catalog_pct == 0:
            continue  # skip brands with no gummies (e.g. beverage-only)

        # AOV from brand's actual median gummy price
        prices = bs["gummy_prices"]
        med_price = sorted(prices)[len(prices) // 2] if prices else 35

        # ── Conservative estimate ──
        # Cart factor 1.3x, conv 1-3%, catalog-based gummy share
        aov_low = med_price * 1.3
        conv_low = max(0.01, 0.03 - (bounce * 0.02))
        gummy_pct_low = gummy_catalog_pct
        rev_low = visits * gummy_pct_low * conv_low * aov_low

        # ── Calibrated estimate ──
        # Cart factor 2.5x, conv 1-5%, gummy traffic multiplier 1.8x
        # (calibrated to Mood ~$3M/mo known gummy revenue)
        aov_high = med_price * 2.5
        conv_high = max(0.01, 0.05 - (bounce * 0.04))
        gummy_pct_high = min(1.0, gummy_catalog_pct * 1.8)
        rev_high = visits * gummy_pct_high * conv_high * aov_high

        dtc_revenue[brand] = {
            "revenue_low": rev_low,
            "revenue_high": rev_high,
            "revenue": (rev_low + rev_high) / 2,  # midpoint
            "visits": visits,
            "gummy_pct_low": gummy_pct_low,
            "gummy_pct_high": gummy_pct_high,
            "conv_low": conv_low,
            "conv_high": conv_high,
            "bounce": bounce,
            "med_price": med_price,
            "aov_low": aov_low,
            "aov_high": aov_high,
            "domain": domain,
        }

    dtc_total_low = sum(d["revenue_low"] for d in dtc_revenue.values())
    dtc_total_high = sum(d["revenue_high"] for d in dtc_revenue.values())
    dtc_total = (dtc_total_low + dtc_total_high) / 2

    # Dispensary: rough estimate from product count
    disp_products = sum(1 for p in products if p["source"] == "Dispensary")
    disp_avg_price = 25  # typical dispensary gummy price
    # Rough: each unique product represents ~100 dispensary sales/mo
    disp_total = disp_products * 100 * disp_avg_price if disp_products > 0 else 0

    # ── Industry TAM model ───────────────────────────────────────────────
    # Uses published data from BDSA, Headset, and Brightfield to build
    # a top-down market size estimate for gummies and functional gummies.
    #
    # Sources:
    #   BDSA (2025): US regulated edibles = $4.3B, intox hemp edibles = 27% of $21.8B
    #   Headset (2022-2026): Gummies = 72-77% of edible sales
    #   BDSA Q1 2025: Candy = 79% of edibles, gummies = 85% of candy → ~67%
    #   Brightfield (2025): Hemp-derived THC gummies = $1.34B
    #   Our product data: functional share by channel

    # Compute our functional share by source
    func_by_source = defaultdict(lambda: {"total": 0, "functional": 0, "func_revenue": 0, "total_revenue": 0})
    for p in products:
        src = p["source"]
        func_by_source[src]["total"] += 1
        rev = (p.get("price") or 0) * (p.get("soldPastMonth") or 0)
        func_by_source[src]["total_revenue"] += rev
        if p.get("effects"):
            func_by_source[src]["functional"] += 1
            func_by_source[src]["func_revenue"] += rev

    # Functional share by channel (from our data where available, estimated otherwise)
    # Amazon: our data. DTC: our data. Dispensary: estimated lower (more recreational)
    amz_func_share = (func_by_source["Amazon"]["functional"]
                      / max(func_by_source["Amazon"]["total"], 1))
    amz_func_rev_share = (func_by_source["Amazon"]["func_revenue"]
                          / max(func_by_source["Amazon"]["total_revenue"], 1))
    dtc_func_share = (func_by_source["DTC"]["functional"]
                      / max(func_by_source["DTC"]["total"], 1))
    # Dispensary: lower functional share — more recreational use
    # Sleep/CBN gummies are top sellers but most dispensary gummies are recreational
    # Estimate 25-35% functional based on Headset SKU-level signals
    disp_func_share = 0.30

    # --- Channel sizes (annual, USD) ---
    # Regulated dispensary gummies
    bdsa_regulated_edibles = 4.3e9     # BDSA 2025: US regulated edibles
    headset_gummy_share = 0.73         # Headset: gummies = 72-77% of edibles (midpoint)
    tam_dispensary_gummies = bdsa_regulated_edibles * headset_gummy_share  # ~$3.14B

    # Hemp-derived THC gummies — use Brightfield as conservative
    bf_hemp_gummies = None
    if hasattr(compute_market_size, '_brightfield_gummies'):
        bf_hemp_gummies = compute_market_size._brightfield_gummies
    tam_hemp_gummies_conservative = bf_hemp_gummies or 1.34e9  # Brightfield direct measurement
    # High estimate: 2x Brightfield to account for undercounting in unregulated channels.
    # BDSA's $21.8B total hemp figure (which implies $3.9B gummies) is controversial —
    # it would mean unregulated hemp is 4x the entire regulated dispensary market.
    # 2x Brightfield is a more defensible upper bound.
    tam_hemp_gummies_aggressive = tam_hemp_gummies_conservative * 2.0

    # Amazon hemp/CBD gummies (our measured + extrapolated, annualized)
    tam_amazon = (amz_measured + amz_extrapolated) * 12

    # Total gummy TAM range
    tam_total_low = tam_dispensary_gummies + tam_hemp_gummies_conservative + tam_amazon
    tam_total_high = tam_dispensary_gummies + tam_hemp_gummies_aggressive + tam_amazon

    # Functional gummy TAM — apply per-channel functional shares
    func_tam_low = (tam_dispensary_gummies * disp_func_share
                    + tam_hemp_gummies_conservative * dtc_func_share
                    + tam_amazon * amz_func_rev_share)
    func_tam_high = (tam_dispensary_gummies * disp_func_share
                     + tam_hemp_gummies_aggressive * dtc_func_share
                     + tam_amazon * amz_func_rev_share)

    return {
        "amazon_measured": amz_measured,
        "amazon_extrapolated": amz_extrapolated,
        "amazon_total": amz_measured + amz_extrapolated,
        "amazon_measured_pct": amz_measured_products / max(amz_total, 1),
        "amazon_measured_products": amz_measured_products,
        "dtc_estimated": dtc_total,
        "dtc_estimated_low": dtc_total_low,
        "dtc_estimated_high": dtc_total_high,
        "dtc_brand_revenues": dtc_revenue,
        "dispensary_estimated": disp_total,
        "total_tracked": amz_measured + amz_extrapolated + dtc_total + disp_total,
        "total_tracked_low": amz_measured + amz_extrapolated + dtc_total_low + disp_total,
        "total_tracked_high": amz_measured + amz_extrapolated + dtc_total_high + disp_total,
        "traffic_brands": len(traffic),
        "dtc_brands_with_traffic": len(dtc_revenue),
        "dtc_brands_count": len(dtc_brands_in_products),
        "dtc_products_count": dtc_products_count,
        # Industry TAM
        "tam_dispensary_gummies": tam_dispensary_gummies,
        "tam_hemp_gummies_low": tam_hemp_gummies_conservative,
        "tam_hemp_gummies_high": tam_hemp_gummies_aggressive,
        "tam_amazon": tam_amazon,
        "tam_total_low": tam_total_low,
        "tam_total_high": tam_total_high,
        "func_tam_low": func_tam_low,
        "func_tam_high": func_tam_high,
        # Functional shares used
        "func_share_amazon": amz_func_rev_share,
        "func_share_dtc": dtc_func_share,
        "func_share_dispensary": disp_func_share,
    }


# ── Charts ────────────────────────────────────────────────────────────────────
def chart_form_factor_by_marketplace(products):
    sources = sorted(set(p["source"] for p in products),
                     key=lambda s: list(COLORS.keys()).index(s) if s in COLORS else 99)
    ff_counts = {s: Counter() for s in sources}
    for p in products:
        ff = p.get("formFactor") or "Other"
        ff_counts[p["source"]][ff] += 1
    all_ff = Counter()
    for c in ff_counts.values():
        all_ff.update(c)
    top_ff = [ff for ff, _ in all_ff.most_common(10)]

    fig = go.Figure()
    for ff in top_ff:
        fig.add_trace(go.Bar(name=ff, x=sources, y=[ff_counts[s][ff] for s in sources]))
    fig.update_layout(barmode="stack", title="Product Form Factor by Marketplace",
                      xaxis_title="Marketplace", yaxis_title="Products", height=450, template="plotly_white")
    return fig


def chart_cannabinoid_popularity(products):
    sources = sorted(set(p["source"] for p in products),
                     key=lambda s: list(COLORS.keys()).index(s) if s in COLORS else 99)
    cb_counts = {s: Counter() for s in sources}
    for p in products:
        for cb in p.get("cannabinoids", []):
            cb_counts[p["source"]][cb] += 1
    all_cb = Counter()
    for c in cb_counts.values():
        all_cb.update(c)
    top_cb = [cb for cb, _ in all_cb.most_common(12)]
    top_cb.reverse()

    fig = go.Figure()
    for source in sources:
        fig.add_trace(go.Bar(
            name=source, y=top_cb,
            x=[cb_counts[source].get(cb, 0) for cb in top_cb],
            orientation="h", marker_color=COLORS.get(source, "#999"),
        ))
    fig.update_layout(barmode="stack", title="Cannabinoid Popularity",
                      xaxis_title="Number of Products", height=500, template="plotly_white")
    return fig


def chart_effect_popularity(products):
    sources = sorted(set(p["source"] for p in products),
                     key=lambda s: list(COLORS.keys()).index(s) if s in COLORS else 99)
    eff_counts = {s: Counter() for s in sources}
    for p in products:
        for eff in p.get("effects", []):
            eff_counts[p["source"]][eff] += 1
    all_eff = Counter()
    for c in eff_counts.values():
        all_eff.update(c)
    top_eff = [e for e, _ in all_eff.most_common(10)]
    top_eff.reverse()

    fig = go.Figure()
    for source in sources:
        fig.add_trace(go.Bar(
            name=source, y=top_eff,
            x=[eff_counts[source].get(e, 0) for e in top_eff],
            orientation="h", marker_color=COLORS.get(source, "#999"),
        ))
    fig.update_layout(barmode="stack", title="Functional Category Breakdown",
                      xaxis_title="Number of Products", height=500, template="plotly_white")
    return fig


def chart_functional_revenue(products):
    """Revenue breakdown by functional category (Amazon only — has sold data)."""
    eff_rev = defaultdict(float)
    eff_count = defaultdict(int)
    general_rev = 0
    general_count = 0
    for p in products:
        if not (p.get("soldPastMonth") and p.get("price")):
            continue
        rev = p["price"] * p["soldPastMonth"]
        effects = p.get("effects", [])
        if effects:
            for eff in effects:
                eff_rev[eff] += rev
                eff_count[eff] += 1
        else:
            general_rev += rev
            general_count += 1

    if not eff_rev and not general_rev:
        fig = go.Figure()
        fig.add_annotation(text="No revenue data available", showarrow=False, font=dict(size=14))
        fig.update_layout(height=400, template="plotly_white", title="Revenue by Functional Category")
        return fig

    # Sort by revenue
    items = sorted(eff_rev.items(), key=lambda x: -x[1])
    if general_rev:
        items.append(("General / Unclassified", general_rev))

    labels = [e for e, _ in items]
    values = [r for _, r in items]
    colors = ["#2563EB", "#DC2626", "#059669", "#D97706", "#7C3AED",
              "#DB2777", "#0891B2", "#65A30D", "#EA580C", "#4F46E5",
              "#6B7280"]  # last is gray for General

    labels.reverse()
    values.reverse()

    fig = go.Figure(go.Bar(
        y=labels, x=values, orientation="h",
        marker_color=[colors[i % len(colors)] for i in range(len(labels))],
        text=[f"${r/1000:,.0f}K" for r in values], textposition="outside",
    ))
    total = sum(values)
    fig.update_layout(
        title=f"Est. Monthly Revenue by Functional Category (Total: ${total/1e3:,.0f}K)<br>"
              f"<sup>Amazon only — products may appear in multiple categories</sup>",
        xaxis_title="Est. Monthly Revenue ($)", height=500, template="plotly_white",
        xaxis=dict(tickprefix="$", tickformat=",", range=[0, max(values) * 1.25] if values else None),
    )
    return fig


def chart_functional_share(products):
    """Pie chart: functional vs general products."""
    functional = sum(1 for p in products if p.get("effects"))
    general = sum(1 for p in products if not p.get("effects"))
    fig = go.Figure(go.Pie(
        labels=["Functional", "General"],
        values=[functional, general],
        marker=dict(colors=["#2563EB", "#E5E7EB"]),
        textinfo="label+percent", textfont_size=14,
        hole=0.4,
    ))
    fig.update_layout(
        title=f"Functional vs General ({functional}/{functional+general} functional)",
        height=400, template="plotly_white",
        showlegend=False,
    )
    return fig


def chart_price_distribution(products):
    sources = sorted(set(p["source"] for p in products),
                     key=lambda s: list(COLORS.keys()).index(s) if s in COLORS else 99)
    fig = go.Figure()
    for source in sources:
        prices = [p["price"] for p in products if p["source"] == source and p.get("price")]
        if len(prices) < 5:
            continue
        fig.add_trace(go.Box(
            y=prices, name=source, marker_color=COLORS.get(source, "#999"),
        ))
    fig.update_layout(title="Price Distribution by Channel",
                      yaxis_title="Price ($)", height=450, template="plotly_white")
    return fig


def chart_top_brands(products):
    brand_counts = Counter()
    brand_sources = defaultdict(lambda: Counter())
    for p in products:
        brand = p.get("brand") or "Unknown"
        brand_counts[brand] += 1
        brand_sources[brand][p["source"]] += 1
    top = [b for b, _ in brand_counts.most_common(25) if b != "Unknown"]
    top.reverse()

    sources = sorted(set(p["source"] for p in products),
                     key=lambda s: list(COLORS.keys()).index(s) if s in COLORS else 99)
    fig = go.Figure()
    for source in sources:
        fig.add_trace(go.Bar(
            name=source, y=top,
            x=[brand_sources[b].get(source, 0) for b in top],
            orientation="h", marker_color=COLORS.get(source, "#999"),
        ))
    fig.update_layout(barmode="stack", title="Top 25 Brands by Product Count",
                      xaxis_title="Products", height=700, template="plotly_white")
    return fig


def chart_brand_map(products):
    brands = defaultdict(lambda: {"prices": [], "sold": [], "reviews": [], "count": 0, "sources": set()})
    for p in products:
        brand = p.get("brand") or "Unknown"
        if brand == "Unknown":
            continue
        brands[brand]["count"] += 1
        brands[brand]["sources"].add(p["source"])
        if p.get("price"):
            brands[brand]["prices"].append(p["price"])
        if p.get("soldPastMonth"):
            brands[brand]["sold"].append(p["soldPastMonth"])
        if p.get("reviewCount"):
            brands[brand]["reviews"].append(p["reviewCount"])

    def demand_score(d):
        if d["sold"]:
            return sum(d["sold"])
        if d["reviews"]:
            return max(d["reviews"])
        return 0

    def demand_label(d):
        if d["sold"]:
            return f"Sold/mo: {sum(d['sold']):,}"
        if d["reviews"]:
            return f"Max reviews: {max(d['reviews']):,}"
        return "No demand data"

    # Count products per source for filtering small samples
    source_counts = Counter(p["source"] for p in products)

    fig = go.Figure()
    for source in COLORS:
        if source_counts.get(source, 0) < 5:
            continue
        subset = [(b, d) for b, d in brands.items()
                  if source in d["sources"] and demand_score(d) > 0]
        if not subset:
            continue
        fig.add_trace(go.Scatter(
            x=[sum(d["prices"]) / len(d["prices"]) if d["prices"] else 0 for _, d in subset],
            y=[demand_score(d) for _, d in subset],
            mode="markers+text",
            text=[b if demand_score(d) > 500 else "" for b, d in subset],
            textposition="top center",
            name=source,
            marker=dict(
                size=[min(max(d["count"] * 3, 8), 40) for _, d in subset],
                color=COLORS[source], opacity=0.7,
            ),
            hovertext=[f"<b>{b}</b><br>{d['count']} products<br>"
                       f"Avg price: ${sum(d['prices'])/len(d['prices']):.2f}<br>"
                       f"{demand_label(d)}"
                       for b, d in subset],
            hoverinfo="text",
        ))
    fig.update_layout(
        title="Brand Market Map — Price vs Demand",
        xaxis_title="Avg Price ($)", yaxis_title="Demand (sold/mo or reviews, log)",
        xaxis=dict(range=[0, 200]),
        yaxis=dict(type="log"), height=550, template="plotly_white",
    )
    return fig


def chart_market_map(products):
    """Marimekko: form factor columns (width=revenue), brand stacks."""
    ff_brand_rev = defaultdict(lambda: defaultdict(float))
    ff_totals = defaultdict(float)
    for p in products:
        if not (p.get("soldPastMonth") and p["soldPastMonth"] > 0 and p.get("price")):
            continue
        rev = p["price"] * p["soldPastMonth"]
        ff = p.get("formFactor") or "Other"
        brand = p.get("brand") or "Unknown"
        ff_brand_rev[ff][brand] += rev
        ff_totals[ff] += rev

    if not ff_totals:
        fig = go.Figure()
        fig.add_annotation(text="No revenue data available (Amazon only)", showarrow=False, font=dict(size=14))
        fig.update_layout(height=400, template="plotly_white", title="Estimated Revenue by Form Factor")
        return fig

    sorted_ff = sorted(ff_totals.keys(), key=lambda k: -ff_totals[k])[:8]
    total_rev = sum(ff_totals[ff] for ff in sorted_ff)
    MAX_BRANDS = 6

    ff_stacks = {}
    all_brand_names = set()
    for ff in sorted_ff:
        top = sorted(ff_brand_rev[ff].items(), key=lambda x: -x[1])[:MAX_BRANDS]
        rest = sum(r for _, r in sorted(ff_brand_rev[ff].items(), key=lambda x: -x[1])[MAX_BRANDS:])
        stack = list(top)
        if rest > 0:
            stack.append(("Other Brands", rest))
        ff_stacks[ff] = stack
        for b, _ in stack:
            all_brand_names.add(b)

    palette = ["#2563EB", "#DC2626", "#059669", "#D97706", "#7C3AED",
               "#DB2777", "#0891B2", "#65A30D", "#EA580C", "#4F46E5",
               "#BE185D", "#0D9488", "#B45309", "#7C2D12", "#6D28D9"]
    brand_color = {}
    ci = 0
    for b in sorted(all_brand_names):
        if b == "Other Brands":
            brand_color[b] = "#E5E7EB"
        else:
            brand_color[b] = palette[ci % len(palette)]
            ci += 1

    MIN_WIDTH_PCT = 8
    raw_widths = {ff: math.sqrt(ff_totals[ff]) for ff in sorted_ff}
    raw_sum = sum(raw_widths.values())
    widths = [(ff, max(100 * raw_widths[ff] / raw_sum, MIN_WIDTH_PCT)) for ff in sorted_ff]
    w_sum = sum(w for _, w in widths)
    widths = [(ff, w * 100 / w_sum) for ff, w in widths]
    gap = 1.0
    total_gap = gap * (len(widths) - 1)
    scale = (100 - total_gap) / 100
    widths = [(ff, w * scale) for ff, w in widths]

    x_starts = []
    x = 0
    for ff, w in widths:
        x_starts.append(x)
        x += w + gap

    fig = go.Figure()
    for col_idx, (ff, col_width) in enumerate(widths):
        x_center = x_starts[col_idx] + col_width / 2
        ff_rev = ff_totals[ff]
        stack = ff_stacks[ff]
        y_bottom = 0
        for brand, rev in reversed(stack):
            pct = 100 * rev / ff_rev
            fig.add_trace(go.Bar(
                x=[x_center], y=[pct], width=col_width, base=y_bottom,
                marker_color=brand_color.get(brand, "#999"),
                marker_line=dict(color="white", width=1.5),
                showlegend=False,
                hovertext=f"<b>{brand}</b><br>{ff}<br>${rev:,.0f}/mo ({pct:.1f}%)",
                hoverinfo="text",
            ))
            if pct >= 6 and col_width >= 5:
                label = f"<b>{brand}</b><br>${rev/1000:,.0f}K" if pct >= 15 else brand[:15]
                fig.add_annotation(
                    x=x_center, y=y_bottom + pct / 2, text=label, showarrow=False,
                    font=dict(size=10, color="white" if brand != "Other Brands" else "#555"),
                )
            y_bottom += pct

    fig.update_layout(
        xaxis=dict(
            tickmode="array",
            tickvals=[x_starts[i] + widths[i][1] / 2 for i in range(len(widths))],
            ticktext=[f"<b>{ff}</b><br>${ff_totals[ff]/1000:,.0f}K ({100*ff_totals[ff]/total_rev:.0f}%)" for ff, _ in widths],
            range=[-1, 101], showgrid=False,
        ),
        yaxis=dict(title="Brand Share (%)", range=[0, 105], ticksuffix="%"),
        barmode="overlay",
        title=f"Brand Revenue Share (Total: ${total_rev/1e6:.1f}M/mo est.)<br>"
              f"<sup>Amazon only (&lt;1% of TAM) — shown because it's our only measured channel</sup>",
        height=600, template="plotly_white", margin=dict(t=80, b=100),
    )
    return fig


def chart_functional_market_map(products):
    """Marimekko: functional category columns (width=revenue), brand stacks."""
    cat_brand_rev = defaultdict(lambda: defaultdict(float))
    cat_totals = defaultdict(float)
    for p in products:
        if not (p.get("soldPastMonth") and p["soldPastMonth"] > 0 and p.get("price")):
            continue
        rev = p["price"] * p["soldPastMonth"]
        brand = p.get("brand") or "Unknown"
        effects = p.get("effects", [])
        if not effects:
            effects = ["General"]
        for eff in effects:
            cat_brand_rev[eff][brand] += rev
            cat_totals[eff] += rev

    if not cat_totals:
        fig = go.Figure()
        fig.add_annotation(text="No revenue data available", showarrow=False, font=dict(size=14))
        fig.update_layout(height=400, template="plotly_white",
                          title="Functional Category Market Map")
        return fig

    sorted_cats = sorted(cat_totals.keys(), key=lambda k: -cat_totals[k])[:10]
    total_rev = sum(cat_totals[c] for c in sorted_cats)
    MAX_BRANDS = 5

    cat_stacks = {}
    all_brand_names = set()
    for cat in sorted_cats:
        top = sorted(cat_brand_rev[cat].items(), key=lambda x: -x[1])[:MAX_BRANDS]
        rest = sum(r for _, r in sorted(cat_brand_rev[cat].items(), key=lambda x: -x[1])[MAX_BRANDS:])
        stack = list(top)
        if rest > 0:
            stack.append(("Other Brands", rest))
        cat_stacks[cat] = stack
        for b, _ in stack:
            all_brand_names.add(b)

    palette = ["#2563EB", "#DC2626", "#059669", "#D97706", "#7C3AED",
               "#DB2777", "#0891B2", "#65A30D", "#EA580C", "#4F46E5",
               "#BE185D", "#0D9488", "#B45309", "#7C2D12", "#6D28D9"]
    brand_color = {}
    ci = 0
    for b in sorted(all_brand_names):
        if b == "Other Brands":
            brand_color[b] = "#E5E7EB"
        else:
            brand_color[b] = palette[ci % len(palette)]
            ci += 1

    MIN_WIDTH_PCT = 6
    raw_widths = {c: math.sqrt(cat_totals[c]) for c in sorted_cats}
    raw_sum = sum(raw_widths.values())
    widths = [(c, max(100 * raw_widths[c] / raw_sum, MIN_WIDTH_PCT)) for c in sorted_cats]
    w_sum = sum(w for _, w in widths)
    widths = [(c, w * 100 / w_sum) for c, w in widths]
    gap = 1.0
    total_gap = gap * (len(widths) - 1)
    scale = (100 - total_gap) / 100
    widths = [(c, w * scale) for c, w in widths]

    x_starts = []
    x = 0
    for c, w in widths:
        x_starts.append(x)
        x += w + gap

    fig = go.Figure()
    for col_idx, (cat, col_width) in enumerate(widths):
        x_center = x_starts[col_idx] + col_width / 2
        cat_rev = cat_totals[cat]
        stack = cat_stacks[cat]
        y_bottom = 0
        for brand, rev in reversed(stack):
            pct = 100 * rev / cat_rev
            fig.add_trace(go.Bar(
                x=[x_center], y=[pct], width=col_width, base=y_bottom,
                marker_color=brand_color.get(brand, "#999"),
                marker_line=dict(color="white", width=1.5),
                showlegend=False,
                hovertext=f"<b>{brand}</b><br>{cat}<br>${rev:,.0f}/mo ({pct:.1f}%)",
                hoverinfo="text",
            ))
            if pct >= 8 and col_width >= 5:
                label = f"<b>{brand}</b><br>${rev/1000:,.0f}K" if pct >= 18 else brand[:12]
                fig.add_annotation(
                    x=x_center, y=y_bottom + pct / 2, text=label, showarrow=False,
                    font=dict(size=9, color="white" if brand != "Other Brands" else "#555"),
                )
            y_bottom += pct

    fig.update_layout(
        xaxis=dict(
            tickmode="array",
            tickvals=[x_starts[i] + widths[i][1] / 2 for i in range(len(widths))],
            ticktext=[f"<b>{cat}</b><br>${cat_totals[cat]/1000:,.0f}K ({100*cat_totals[cat]/total_rev:.0f}%)"
                      for cat, _ in widths],
            range=[-1, 101], showgrid=False,
        ),
        yaxis=dict(title="Brand Share (%)", range=[0, 105], ticksuffix="%"),
        barmode="overlay",
        title=f"Functional Category Market Map (Total: ${total_rev/1e3:,.0f}K/mo est.)<br>"
              f"<sup>Column width ∝ category revenue — Amazon measured data</sup>",
        height=600, template="plotly_white", margin=dict(t=80, b=120),
    )
    return fig


# Brightfield consumer survey: "Hemp THC Gummies Desired Effects"
# Source: Brightfield Group, Nov 2025
BRIGHTFIELD_DESIRED_EFFECTS = [
    ("Relax", 63.7),
    ("Sleep", 49.4),
    ("Emotional relief", 48.6),
    ("Physical relief", 42.0),
    ("General well-being", 34.5),
    ("Fun/celebrations", 25.8),
    ("Focus", 25.6),
    ("Stimulate appetite", 23.3),
    ("Energy", 21.0),
    ("Spark creativity", 18.8),
    ("Enhance intimacy", 14.1),
    ("Athletic/gym performance", 6.8),
]

# Mapping from Brightfield desired effects → our EFFECT_RULES categories
BF_TO_OUR_EFFECTS = {
    "Relax": "Relaxation",
    "Sleep": "Sleep",
    "Emotional relief": "Relaxation",  # stress/anxiety maps to our Relaxation
    "Physical relief": "Pain",
    "General well-being": None,  # no direct match
    "Fun/celebrations": "Social",
    "Focus": "Focus",
    "Stimulate appetite": None,
    "Energy": "Focus",  # energy maps to our Focus category
    "Spark creativity": None,
    "Enhance intimacy": "Intimacy",
    "Athletic/gym performance": None,
}


def chart_demand_vs_supply(products):
    """Compare Brightfield consumer demand vs product supply (by count and by revenue)."""
    total_products = len(products)
    if not total_products:
        fig = go.Figure()
        fig.add_annotation(text="No product data", showarrow=False, font=dict(size=14))
        fig.update_layout(height=400, template="plotly_white")
        return fig

    # Count products per effect category
    eff_counts = Counter()
    eff_revenue = defaultdict(float)
    total_revenue = 0
    for p in products:
        rev = (p.get("price") or 0) * (p.get("soldPastMonth") or 0)
        total_revenue += rev
        for eff in p.get("effects", []):
            eff_counts[eff] += 1
            eff_revenue[eff] += rev

    # Build comparison data
    bf_labels = []
    bf_demand = []
    supply_count = []
    supply_revenue = []

    for bf_name, bf_pct in BRIGHTFIELD_DESIRED_EFFECTS:
        our_cat = BF_TO_OUR_EFFECTS.get(bf_name)
        count_pct = (eff_counts.get(our_cat, 0) / total_products * 100) if our_cat else 0
        rev_pct = (eff_revenue.get(our_cat, 0) / total_revenue * 100) if (our_cat and total_revenue) else 0
        bf_labels.append(bf_name)
        bf_demand.append(bf_pct)
        supply_count.append(round(count_pct, 1))
        supply_revenue.append(round(rev_pct, 1))

    bf_labels.reverse()
    bf_demand.reverse()
    supply_count.reverse()
    supply_revenue.reverse()

    fig = go.Figure()
    # Build hover text with dollar values for revenue bars
    rev_dollars = []
    for bf_name, _ in BRIGHTFIELD_DESIRED_EFFECTS:
        our_cat = BF_TO_OUR_EFFECTS.get(bf_name)
        rev = eff_revenue.get(our_cat, 0) if our_cat else 0
        rev_dollars.append(rev)
    rev_dollars.reverse()

    fig.add_trace(go.Bar(
        y=bf_labels, x=bf_demand, name="Consumer Demand (Brightfield)",
        orientation="h", marker_color="#2563EB",
        text=[f"{v:.0f}%" for v in bf_demand], textposition="outside",
        hovertext=[f"{l}: {v:.1f}% of consumers want this" for l, v in zip(bf_labels, bf_demand)],
        hoverinfo="text",
    ))
    fig.add_trace(go.Bar(
        y=bf_labels, x=supply_revenue, name="Supply by Revenue (Amazon)",
        orientation="h", marker_color="#059669",
        text=[f"{v:.0f}%" for v in supply_revenue], textposition="outside",
        hovertext=[f"{l}: {v:.1f}% of revenue (${d/1e3:,.0f}K/mo)" for l, v, d in zip(bf_labels, supply_revenue, rev_dollars)],
        hoverinfo="text",
    ))
    fig.add_trace(go.Bar(
        y=bf_labels, x=supply_count, name="Supply by Product Count",
        orientation="h", marker_color="#F59E0B", opacity=0.7,
        text=[f"{v:.0f}%" for v in supply_count], textposition="outside",
        hovertext=[f"{l}: {v:.1f}% of products" for l, v in zip(bf_labels, supply_count)],
        hoverinfo="text",
    ))
    fig.update_layout(
        barmode="group",
        title="Consumer Demand vs Product Supply by Desired Effect<br>"
              "<sup>Totals exceed 100% — products serve multiple effects and consumers select multiple needs</sup>",
        xaxis_title="Percentage (%)", xaxis=dict(ticksuffix="%", range=[0, 80]),
        height=600, template="plotly_white",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )
    return fig


def chart_brightfield_market_growth(bf):
    """Line chart: hemp-derived THC product type market size 2020-2030."""
    if not bf:
        fig = go.Figure()
        fig.add_annotation(text="Brightfield data not available", showarrow=False, font=dict(size=14))
        fig.update_layout(height=400, template="plotly_white")
        return fig

    years = bf["product_type_years"]
    sizes = bf["product_type_sizes"]

    # Color palette for product types
    type_colors = {
        "Gummies": "#2563EB", "Drinks": "#059669", "Vapes": "#DC2626",
        "Flower & Pre-Rolls": "#D97706", "Other Edibles": "#7C3AED",
        "Concentrates": "#DB2777", "Tinctures": "#0891B2",
        "Capsules": "#65A30D", "Beauty & Topicals": "#EA580C",
    }

    # Sort by 2025 value (index 5) descending
    sorted_types = sorted(sizes.items(),
                          key=lambda x: x[1][5] if len(x[1]) > 5 and x[1][5] else 0,
                          reverse=True)

    fig = go.Figure()
    for name, vals in sorted_types:
        # Convert to billions
        y_vals = [v / 1e9 if v is not None else None for v in vals]
        val_2025 = vals[5] / 1e9 if len(vals) > 5 and vals[5] else 0
        hover_texts = [f"<b>{name}</b><br>{yr}: ${v:.2f}B" if v is not None else ""
                       for yr, v in zip(years, y_vals)]
        fig.add_trace(go.Scatter(
            x=years, y=y_vals, name=f"{name} (${val_2025:.1f}B)",
            mode="lines+markers",
            line=dict(color=type_colors.get(name, "#999"), width=3 if name == "Gummies" else 2),
            marker=dict(size=6 if name == "Gummies" else 4),
            hovertext=hover_texts, hoverinfo="text",
        ))

    total_2025 = sum(v[5] for v in sizes.values() if v and len(v) > 5 and v[5]) / 1e9
    fig.add_vline(x=2025, line_dash="dot", line_color="#999", annotation_text="Current")
    fig.update_layout(
        title=f"Hemp-Derived THC Market by Product Type (Total 2025: ${total_2025:.1f}B)<br>"
              f"<sup>Source: Brightfield Group — values in USD billions</sup>",
        xaxis_title="Year", yaxis_title="Market Size ($B)",
        yaxis=dict(tickprefix="$", ticksuffix="B", tickformat=".1f"),
        height=500, template="plotly_white",
        legend=dict(font=dict(size=11)),
    )
    return fig


def chart_brightfield_cannabinoid_trends(bf):
    """Line chart: cannabinoid market share trends 2022-2025."""
    if not bf:
        fig = go.Figure()
        fig.add_annotation(text="Brightfield data not available", showarrow=False, font=dict(size=14))
        fig.update_layout(height=400, template="plotly_white")
        return fig

    years = bf["cannabinoid_years"]
    shares = bf["cannabinoid_shares"]

    cb_colors = {
        "Delta-9": "#2563EB", "Delta-8": "#DC2626", "THCA": "#059669",
        "THCP": "#D97706", "HHC": "#7C3AED", "Others": "#9CA3AF",
        "Delta-10": "#DB2777", "THCV": "#0891B2", "THCH": "#65A30D",
        "THCJD": "#EA580C",
    }

    # Sort by 2025 share descending
    sorted_cbs = sorted(shares.items(),
                        key=lambda x: x[1][3] if len(x[1]) > 3 and x[1][3] else 0,
                        reverse=True)

    fig = go.Figure()
    for name, vals in sorted_cbs:
        share_2025 = vals[3] if len(vals) > 3 and vals[3] else 0
        fig.add_trace(go.Scatter(
            x=years, y=vals, name=f"{name} ({share_2025:.1f}%)",
            mode="lines+markers",
            line=dict(color=cb_colors.get(name, "#999"), width=3 if share_2025 > 15 else 2),
            marker=dict(size=8 if share_2025 > 15 else 5),
        ))

    fig.update_layout(
        title="Cannabinoid Market Share Trends (Hemp-Derived THC)<br>"
              "<sup>Source: Brightfield Group — Delta-9 rising, Delta-8 declining, THCA emerging</sup>",
        xaxis_title="Year", yaxis_title="Market Share (%)",
        yaxis=dict(ticksuffix="%", range=[0, 60]),
        xaxis=dict(dtick=1),
        height=500, template="plotly_white",
        legend=dict(font=dict(size=11)),
    )
    return fig


def chart_brightfield_company_share(bf):
    """Horizontal bar: top companies by 2025 brand awareness with 2022 comparison."""
    if not bf:
        fig = go.Figure()
        fig.add_annotation(text="Brightfield data not available", showarrow=False, font=dict(size=14))
        fig.update_layout(height=400, template="plotly_white")
        return fig

    companies = bf["company_shares"]

    # Get companies with 2025 data, sorted by share
    co_2025 = [(name, vals[3]) for name, vals in companies.items()
               if len(vals) > 3 and vals[3] is not None]
    co_2025.sort(key=lambda x: x[1], reverse=True)
    top_20 = co_2025[:20]
    top_20.reverse()  # bottom-up for horizontal bar

    names = [n for n, _ in top_20]
    shares_2025 = [s for _, s in top_20]

    # Get 2022 shares for comparison
    shares_2022 = []
    for name in names:
        vals = companies.get(name, [])
        shares_2022.append(vals[0] if len(vals) > 0 and vals[0] is not None else 0)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=names, x=shares_2022, name="2022", orientation="h",
        marker_color="#CBD5E1", text=[f"{s:.1f}%" if s else "" for s in shares_2022],
        textposition="outside",
    ))
    fig.add_trace(go.Bar(
        y=names, x=shares_2025, name="2025", orientation="h",
        marker_color="#2563EB", text=[f"{s:.2f}%" for s in shares_2025],
        textposition="outside",
    ))

    fig.update_layout(
        barmode="group",
        title="Brand Awareness — Hemp-Derived THC<br>"
              "<sup>Source: Brightfield Group consumer survey — highly fragmented, #1 brand recognized by only 1.6%</sup>",
        xaxis_title="Brand Awareness (%)", xaxis=dict(ticksuffix="%", range=[0, max(shares_2025 + shares_2022) * 1.4]),
        height=650, template="plotly_white",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )
    return fig


def chart_top_revenue(products):
    brand_rev = defaultdict(float)
    for p in products:
        if p.get("soldPastMonth") and p.get("price"):
            brand = p.get("brand") or "Unknown"
            brand_rev[brand] += p["price"] * p["soldPastMonth"]
    if not brand_rev:
        fig = go.Figure()
        fig.add_annotation(text="No revenue data", showarrow=False)
        fig.update_layout(height=400, template="plotly_white")
        return fig

    top = sorted(brand_rev.items(), key=lambda x: -x[1])[:20]
    top.reverse()
    fig = go.Figure(go.Bar(
        y=[b for b, _ in top], x=[r for _, r in top],
        orientation="h", marker_color="#2563EB",
        text=[f"${r/1000:,.0f}K" for _, r in top], textposition="outside",
    ))
    max_rev = max(r for _, r in top) if top else 0
    fig.update_layout(title="Top 20 Brands by Est. Monthly Revenue (Amazon)",
                      xaxis_title="Est. Monthly Revenue ($)", height=600, template="plotly_white",
                      xaxis=dict(tickprefix="$", tickformat=",", range=[0, max_rev * 1.25]))
    return fig


def chart_dtc_revenue(market_size):
    """Horizontal bar chart with conservative/calibrated range for DTC brand revenue."""
    dtc_rev = market_size.get("dtc_brand_revenues", {})
    if not dtc_rev:
        fig = go.Figure()
        fig.add_annotation(text="No DTC traffic data available", showarrow=False, font=dict(size=14))
        fig.update_layout(height=400, template="plotly_white",
                          title="DTC Brand Revenue Estimates (traffic-based)")
        return fig

    # Sort by high estimate
    top = sorted(dtc_rev.items(), key=lambda x: -x[1]["revenue_high"])[:25]
    top.reverse()

    # Conservative bars (light purple)
    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=[b for b, _ in top],
        x=[d["revenue_low"] for _, d in top],
        orientation="h", marker_color="#D6BCFA", name="Conservative",
        text=[f"${d['revenue_low']/1000:,.0f}K" for _, d in top],
        textposition="outside",
        hovertext=[
            f"<b>{b} — Conservative</b><br>"
            f"Visits: {d['visits']:,}/mo<br>"
            f"Gummy share: {d['gummy_pct_low']:.0%} (catalog)<br>"
            f"Conv: {d['conv_low']:.1%} (1-3% range)<br>"
            f"AOV: ${d['aov_low']:.0f} (price × 1.3)<br>"
            f"<b>${d['revenue_low']:,.0f}/mo</b>"
            for b, d in top
        ],
        hoverinfo="text",
    ))

    # Calibrated bars (dark purple)
    fig.add_trace(go.Bar(
        y=[b for b, _ in top],
        x=[d["revenue_high"] for _, d in top],
        orientation="h", marker_color="#7C3AED", name="Calibrated",
        text=[f"${d['revenue_high']/1000:,.0f}K" for _, d in top],
        textposition="outside",
        hovertext=[
            f"<b>{b} — Calibrated</b><br>"
            f"Visits: {d['visits']:,}/mo<br>"
            f"Gummy traffic: {d['gummy_pct_high']:.0%} (catalog × 1.8)<br>"
            f"Conv: {d['conv_high']:.1%} (1-5% range)<br>"
            f"AOV: ${d['aov_high']:.0f} (price × 2.5)<br>"
            f"<b>${d['revenue_high']:,.0f}/mo</b>"
            for b, d in top
        ],
        hoverinfo="text",
    ))

    max_rev = max(d["revenue_high"] for _, d in top) if top else 0
    total_lo = sum(d["revenue_low"] for d in dtc_rev.values())
    total_hi = sum(d["revenue_high"] for d in dtc_rev.values())
    fig.update_layout(
        barmode="group",
        title=f"Top DTC Brands by Est. Monthly Gummy Revenue (${total_lo/1e6:.1f}M – ${total_hi/1e6:.1f}M/mo)<br>"
              f"<sup>Conservative (1.3x cart, 1-3% conv) vs Calibrated (2.5x cart, 1-5% conv, 1.8x gummy traffic)</sup>",
        xaxis_title="Est. Monthly Revenue ($)", height=750, template="plotly_white",
        xaxis=dict(tickprefix="$", tickformat=",", range=[0, max_rev * 1.3]),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )
    return fig


def build_dtc_decomposition_html(market_size, brightfield=None):
    """Build an HTML table showing the full DTC revenue decomposition with low/high."""
    dtc_rev = market_size.get("dtc_brand_revenues", {})
    if not dtc_rev:
        return ""

    # Brightfield company market share → implied monthly revenue
    # NOTE: Brightfield data is survey-based, not transaction data
    bf_company_monthly = {}
    bf = brightfield or {}
    bf_total_market = bf.get("total_market_2025")
    bf_companies = bf.get("company_shares", {})
    # Map Brightfield company names → our brand names
    BF_NAME_MAP = {
        "TRĒ House (CBDfx)": "TRĒ House", "Mood": "Mood", "3CHI": "3Chi",
        "CANN": "CANN", "Hometown Hero": "Hometown Hero",
        "Mellow Fellow": "Mellow Fellow", "Urb (LFTD Partners)": "Urb",
        "Delta Extrax": "Delta Extrax", "Crescent": "Crescent Canna",
        "Binoid": "Binoid", "Delta Munchies": "Delta Munchies",
        "Hemp Bombs (Global Widget)": "Hemp Bombs",
        "Cycling Frog": "Cycling Frog", "Koi": "Koi CBD",
        "Exhale Wellness (Cheef Holdings)": "Cheef Botanicals",
        "Five (Medterra)": "Five CBD", "BREZ": "BREZ",
        "Wynk": "WYNK", "Enjoy Hemp": "Enjoy Hemp",
        "Tillman's Tranquils": "Tillmans Tranquils",
        "ELYXR": "Elyxr",
    }
    if bf_total_market and bf_companies:
        for bf_name, shares in bf_companies.items():
            our_name = BF_NAME_MAP.get(bf_name.strip())
            if not our_name:
                continue
            # Use 2025 share (index 3)
            share_2025 = shares[3] if len(shares) > 3 and shares[3] is not None else None
            if share_2025:
                annual = bf_total_market * (share_2025 / 100)
                bf_company_monthly[our_name] = annual / 12

    top = sorted(dtc_rev.items(), key=lambda x: -x[1]["revenue_high"])

    rows = ""
    for brand, d in top:
        if d["revenue_high"] < 1:
            continue
        bf_val = bf_company_monthly.get(brand)
        bf_cell = f"${bf_val:,.0f}" if bf_val else "—"
        bf_style = "color:#059669; font-weight:500;" if bf_val else "color:#d1d5db;"
        rows += f"""<tr>
          <td style="padding:5px 6px; font-weight:500;">{brand}</td>
          <td style="padding:5px 6px; text-align:right;">{d['visits']:,}</td>
          <td style="padding:5px 6px; text-align:right;">{d['gummy_pct_low']:.0%}</td>
          <td style="padding:5px 6px; text-align:right;">{d['gummy_pct_high']:.0%}</td>
          <td style="padding:5px 6px; text-align:right;">{d['bounce']:.0%}</td>
          <td style="padding:5px 6px; text-align:right;">{d['conv_low']:.1%}</td>
          <td style="padding:5px 6px; text-align:right;">{d['conv_high']:.1%}</td>
          <td style="padding:5px 6px; text-align:right;">${d['med_price']:.0f}</td>
          <td style="padding:5px 6px; text-align:right;">${d['aov_low']:.0f}</td>
          <td style="padding:5px 6px; text-align:right;">${d['aov_high']:.0f}</td>
          <td style="padding:5px 6px; text-align:right; color:#9B59B6;">${d['revenue_low']:,.0f}</td>
          <td style="padding:5px 6px; text-align:right; font-weight:600; color:#7C3AED;">${d['revenue_high']:,.0f}</td>
          <td style="padding:5px 6px; text-align:right; {bf_style}">{bf_cell}</td>
        </tr>"""

    total_lo = sum(d["revenue_low"] for d in dtc_rev.values())
    total_hi = sum(d["revenue_high"] for d in dtc_rev.values())
    n_brands = len([d for d in dtc_rev.values() if d["revenue_high"] >= 1])
    total_visits = sum(d["visits"] for d in dtc_rev.values())

    return f"""
    <div style="margin-top:16px;">
      <h4 style="font-size:14px; margin-bottom:8px; color:#2d3748;">Revenue Decomposition by Brand</h4>
      <div style="font-size:11px; color:#718096; margin-bottom:12px; line-height:1.7; padding:12px; background:#f9fafb; border-radius:6px;">

        <div style="font-weight:700; font-size:12px; color:#2d3748; margin-bottom:6px;">Model Formula</div>
        <div style="background:white; padding:8px 12px; border-radius:4px; border:1px solid #e2e8f0; font-family:monospace; font-size:11px; margin-bottom:8px;">
          Est. Gummy Revenue = Monthly Visits &times; Gummy Traffic % &times; Conv. Rate &times; AOV
        </div>

        <div style="display:flex; gap:16px; flex-wrap:wrap; margin-bottom:8px;">
          <div style="flex:1; min-width:280px; padding:8px; background:white; border-radius:4px; border-left:3px solid #D6BCFA;">
            <span style="font-weight:600; color:#9B59B6;">Conservative Model</span><br>
            &bull; Gummy % = catalog share (% of brand's SKUs that are gummies)<br>
            &bull; Conv = max(1%, 3% &minus; bounce &times; 2%) &nbsp;<span style="color:#9ca3af;">range: 1&ndash;3%</span><br>
            &bull; AOV = median gummy price &times; 1.3 cart factor
          </div>
          <div style="flex:1; min-width:280px; padding:8px; background:white; border-radius:4px; border-left:3px solid #7C3AED;">
            <span style="font-weight:600; color:#7C3AED;">Calibrated Model</span><br>
            &bull; Gummy % = catalog share &times; 1.8 (gummies drive outsized traffic)<br>
            &bull; Conv = max(1%, 5% &minus; bounce &times; 4%) &nbsp;<span style="color:#9ca3af;">range: 1&ndash;5%</span><br>
            &bull; AOV = median gummy price &times; 2.5 cart factor (bundles, subscribe &amp; save)
          </div>
        </div>

        <div style="padding:8px; background:white; border-radius:4px; border-left:3px solid #059669; margin-bottom:8px;">
          <span style="font-weight:600; color:#059669;">Brightfield (survey)</span><br>
          Company market share &times; <a href="https://app.brightfieldgroup.com/pages/consumer-hemp-derived-thc" target="_blank" style="color:#059669;">$5.42B total hemp THC market</a> &divide; 12.
          All products (not just gummies). Survey-based &mdash; directional, not transactional.
        </div>

        <div style="font-weight:700; font-size:12px; color:#2d3748; margin-top:12px; margin-bottom:6px;">Calibration &mdash; 3 Known Data Points</div>
        <table style="width:100%; font-size:11px; border-collapse:collapse; background:white; border-radius:4px;">
          <thead><tr style="border-bottom:1px solid #e2e8f0;">
            <th style="padding:4px 8px; text-align:left;">Brand</th>
            <th style="padding:4px 8px; text-align:right;">Actual Gummy Rev/mo</th>
            <th style="padding:4px 8px; text-align:right; color:#9B59B6;">Conservative</th>
            <th style="padding:4px 8px; text-align:right; color:#7C3AED;">Calibrated</th>
            <th style="padding:4px 8px; text-align:right;">Cal/Actual</th>
            <th style="padding:4px 8px; text-align:left;">Source</th>
          </tr></thead>
          <tbody>
            <tr><td style="padding:4px 8px; font-weight:500;">Mood</td>
                <td style="padding:4px 8px; text-align:right;">$2,722,000</td>
                <td style="padding:4px 8px; text-align:right; color:#9B59B6;">$486,133</td>
                <td style="padding:4px 8px; text-align:right; color:#7C3AED;">$2,589,132</td>
                <td style="padding:4px 8px; text-align:right; font-weight:600;">95%</td>
                <td style="padding:4px 8px; font-size:10px;">Internal data &mdash; avg of Oct 2025&ndash;Mar 2026 ($2.38M&ndash;$3.06M range)</td></tr>
            <tr><td style="padding:4px 8px; font-weight:500;">Charlotte&rsquo;s Web</td>
                <td style="padding:4px 8px; text-align:right;">~$1.4M&ndash;$2.1M</td>
                <td style="padding:4px 8px; text-align:right; color:#9B59B6;">$281,193</td>
                <td style="padding:4px 8px; text-align:right; color:#7C3AED;">$1,499,360</td>
                <td style="padding:4px 8px; text-align:right; font-weight:600;">86%</td>
                <td style="padding:4px 8px; font-size:10px;"><a href="https://www.prnewswire.com/news-releases/charlottes-web-reports-2025-fourth-quarter-and-full-year-financial-results-302729592.html" target="_blank" style="color:#2563EB;">FY2025 10-K</a>: $49.9M total, 67% DTC, gummies = largest cat.</td></tr>
            <tr><td style="padding:4px 8px; font-weight:500;">cbdMD</td>
                <td style="padding:4px 8px; text-align:right;">~$345K&ndash;$576K</td>
                <td style="padding:4px 8px; text-align:right; color:#9B59B6;">$89,734</td>
                <td style="padding:4px 8px; text-align:right; color:#7C3AED;">$478,580</td>
                <td style="padding:4px 8px; text-align:right; font-weight:600;">104%</td>
                <td style="padding:4px 8px; font-size:10px;"><a href="https://www.prnewswire.com/news-releases/cbdmd-reports-third-consecutive-year-of-operating-improvement-strengthens-balance-sheet-and-regains-nyse-american-continued-listing-compliance-302647308.html" target="_blank" style="color:#2563EB;">FY2025 10-K</a>: $19.2M total, 72% DTC, gummy % estimated.</td></tr>
          </tbody>
          <tfoot><tr style="border-top:1px solid #e2e8f0; font-weight:600;">
            <td colspan="4" style="padding:4px 8px;">Average calibrated fit</td>
            <td style="padding:4px 8px; text-align:right; color:#374151;">95%</td>
            <td style="padding:4px 8px; font-size:10px; font-weight:400; color:#92400e;">&#9888; In-sample fit only &mdash; model was tuned to these 3 points. Out-of-sample accuracy unknown.</td>
          </tr></tfoot>
        </table>

        <div style="font-weight:700; font-size:12px; color:#2d3748; margin-top:12px; margin-bottom:6px;">Data Sources</div>
        <div style="font-size:10px; line-height:1.6;">
          &bull; <b>Monthly visits:</b> <a href="https://zylalabs.com/api-marketplace/data/site+traffic+api/29" target="_blank" style="color:#2563EB;">Zyla Site Traffic API</a> (SimilarWeb data), Feb 2026 snapshot<br>
          &bull; <b>Gummy catalog %:</b> Scraped from each brand's Shopify <code>/products.json</code> API (Apr 2026)<br>
          &bull; <b>Median gummy price:</b> From scraped product data, variant-level pricing<br>
          &bull; <b>Bounce rate:</b> Zyla/SimilarWeb engagement metrics<br>
          &bull; <b>Brightfield market share:</b> <a href="https://app.brightfieldgroup.com/pages/consumer-hemp-derived-thc" target="_blank" style="color:#2563EB;">Brightfield Group Hemp-Derived THC report</a> (Nov 2025), survey-based<br>
          &bull; <b>Charlotte's Web revenue:</b> <a href="https://www.prnewswire.com/news-releases/charlottes-web-reports-2025-fourth-quarter-and-full-year-financial-results-302729592.html" target="_blank" style="color:#2563EB;">CWEB FY2025 earnings</a> (Mar 2026)<br>
          &bull; <b>cbdMD revenue:</b> <a href="https://www.prnewswire.com/news-releases/cbdmd-reports-third-consecutive-year-of-operating-improvement-strengthens-balance-sheet-and-regains-nyse-american-continued-listing-compliance-302647308.html" target="_blank" style="color:#2563EB;">YCBD FY2025 earnings</a> (Jan 2026)
        </div>

        <div style="font-weight:700; font-size:12px; color:#2d3748; margin-top:12px; margin-bottom:6px;">Limitations</div>
        <div style="font-size:10px; line-height:1.6; color:#92400e;">
          &bull; Traffic data is a single Feb 2026 snapshot &mdash; seasonal variation not captured<br>
          &bull; Gummy catalog % assumes catalog share approximates revenue share; in practice, hero products (gummies) often drive outsized revenue vs catalog weight<br>
          &bull; Conversion rate and AOV are modeled, not observed &mdash; actual values vary by brand maturity, marketing spend, and purchase funnel<br>
          &bull; Model captures DTC website revenue only &mdash; excludes wholesale, retail, marketplace (Amazon/Faire), and dispensary channels<br>
          &bull; Calibrated against 3 brands; accuracy for brands with very different profiles (e.g., beverage-first, dispensary-first) may differ
        </div>
      </div>
      <div style="overflow-x:auto; max-height:600px; overflow-y:auto;">
        <table style="width:100%; font-size:11px; border-collapse:collapse;">
          <thead><tr style="border-bottom:2px solid #e2e8f0; position:sticky; top:0; background:white;">
            <th style="padding:5px 6px; text-align:left;">Brand</th>
            <th style="padding:5px 6px; text-align:right;">Visits/mo</th>
            <th style="padding:5px 6px; text-align:right;">Gum%<br><span style="font-weight:400; color:#9ca3af;">cons.</span></th>
            <th style="padding:5px 6px; text-align:right;">Gum%<br><span style="font-weight:400; color:#9ca3af;">cal.</span></th>
            <th style="padding:5px 6px; text-align:right;">Bounce</th>
            <th style="padding:5px 6px; text-align:right;">Conv<br><span style="font-weight:400; color:#9ca3af;">cons.</span></th>
            <th style="padding:5px 6px; text-align:right;">Conv<br><span style="font-weight:400; color:#9ca3af;">cal.</span></th>
            <th style="padding:5px 6px; text-align:right;">Med$</th>
            <th style="padding:5px 6px; text-align:right;">AOV<br><span style="font-weight:400; color:#9ca3af;">cons.</span></th>
            <th style="padding:5px 6px; text-align:right;">AOV<br><span style="font-weight:400; color:#9ca3af;">cal.</span></th>
            <th style="padding:5px 6px; text-align:right; color:#9B59B6;">Rev/mo<br>Cons.</th>
            <th style="padding:5px 6px; text-align:right; color:#7C3AED;">Rev/mo<br>Cal.</th>
            <th style="padding:5px 6px; text-align:right; color:#059669;">BF Total<br><span style="font-weight:400;">survey</span></th>
          </tr></thead>
          <tbody style="color:#374151;">{rows}
          <tr style="border-top:2px solid #e2e8f0; font-weight:700;">
            <td style="padding:5px 6px;">Total ({n_brands} brands)</td>
            <td style="padding:5px 6px; text-align:right;">{total_visits:,}</td>
            <td colspan="8"></td>
            <td style="padding:5px 6px; text-align:right; color:#9B59B6;">${total_lo:,.0f}</td>
            <td style="padding:5px 6px; text-align:right; color:#7C3AED;">${total_hi:,.0f}</td>
            <td style="padding:5px 6px; text-align:right; color:#059669;">{f"${sum(bf_company_monthly.values()):,.0f}" if bf_company_monthly else "—"}</td>
          </tr></tbody>
        </table>
      </div>
    </div>"""


# ── HTML template ─────────────────────────────────────────────────────────────
def build_html(products, charts, market_size=None, brightfield=None):
    total_products = len(products)
    total_brands = len(set(p.get("brand") for p in products if p.get("brand")))
    marketplace_count = len(set(p["source"] for p in products))
    prices = [p["price"] for p in products if p.get("price")]
    avg_price = f"${sum(prices)/len(prices):.2f}" if prices else "N/A"
    functional_count = sum(1 for p in products if p.get("effects"))
    functional_pct = functional_count / max(total_products, 1)

    # Market size formatting
    ms = market_size or {}
    amz_measured = ms.get("amazon_measured", 0)
    amz_extrapolated = ms.get("amazon_extrapolated", 0)
    amz_total = ms.get("amazon_total", 0)
    amz_pct = ms.get("amazon_measured_pct", 0)
    dtc_est = ms.get("dtc_estimated", 0)
    disp_est = ms.get("dispensary_estimated", 0)
    total_tracked = ms.get("total_tracked", 0)
    total_tracked_lo = ms.get("total_tracked_low", 0)
    total_tracked_hi = ms.get("total_tracked_high", 0)

    def fmt_rev(v):
        if v >= 1e6:
            return f"${v/1e6:.1f}M"
        if v >= 1e3:
            return f"${v/1e3:.0f}K"
        return f"${v:.0f}"

    # Build market context HTML section
    market_context_html = ""
    if ms:
        dtc_brands_n = ms.get('dtc_brands_count', 0)
        dtc_prods_n = ms.get('dtc_products_count', 0)
        dtc_lo = ms.get('dtc_estimated_low', 0)
        dtc_hi = ms.get('dtc_estimated_high', 0)

        amz_products = sum(1 for p in products if p["source"] == "Amazon")
        dtc_products = sum(1 for p in products if p["source"] == "DTC")


        # ── TAM Model card ───────────────────────────────────────────────
        tam_disp = ms.get("tam_dispensary_gummies", 0)
        tam_hemp_lo = ms.get("tam_hemp_gummies_low", 0)
        tam_hemp_hi = ms.get("tam_hemp_gummies_high", 0)
        tam_amz = ms.get("tam_amazon", 0)
        tam_lo = ms.get("tam_total_low", 0)
        tam_hi = ms.get("tam_total_high", 0)
        func_lo = ms.get("func_tam_low", 0)
        func_hi = ms.get("func_tam_high", 0)
        fs_amz = ms.get("func_share_amazon", 0)
        fs_dtc = ms.get("func_share_dtc", 0)
        fs_disp = ms.get("func_share_dispensary", 0)

        def fmt_b(v):
            return f"${v/1e9:.1f}B" if v >= 1e9 else f"${v/1e6:.0f}M"

        tam_coverage_lo = total_tracked_lo * 12 / tam_hi * 100 if tam_hi > 0 else 0
        tam_coverage_hi = total_tracked_hi * 12 / tam_lo * 100 if tam_lo > 0 else 0

        market_context_html = f"""
  <div class="section-header"><h2>Market Size & Context</h2><p>Total addressable market, tracked revenue, and industry benchmarks</p></div>

  <div class="card full-width" style="padding:20px 24px; background:linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%); border-left:4px solid #2563EB;">
    <div style="font-size:14px; color:#1e3a5f; line-height:1.7;">
      <b>Bottom line:</b> We can only directly measure <b>{tam_coverage_lo:.1f}&ndash;{tam_coverage_hi:.1f}%</b> of this market
      ({fmt_rev(total_tracked_lo)}&ndash;{fmt_rev(total_tracked_hi)}/mo across {total_brands} brands on Amazon and DTC).
      The rest &mdash; dispensaries, convenience, gas stations, international &mdash; is modeled from industry reports.
      Supplement gummies targeting the same functional needs are <b>$3.6B&ndash;$6.7B</b> (retail-scanner-measured), suggesting
      significant consumer overlap and competitive pressure.
      <span style="color:#92400e; font-weight:600;">Regulatory change (Nov 2026) may eliminate the hemp-derived channel entirely.</span>
    </div>
  </div>

  <details class="card full-width" style="padding:0; background:linear-gradient(135deg, #fefce8 0%, #fef9c3 100%); border-left:4px solid #ca8a04;">
    <summary style="padding:16px 24px; cursor:pointer; display:flex; align-items:center; gap:12px; list-style:none;">
      <span style="font-size:20px;">&#9888;&#65039;</span>
      <div>
        <span style="font-weight:700; color:#854d0e;">Regulatory Alert — Market Restructuring Ahead</span>
        <span style="font-size:12px; color:#a16207; margin-left:8px;">0.4mg THC cap effective Nov 12, 2026 &mdash; click to expand</span>
      </div>
    </summary>
    <div style="padding:0 24px 20px 60px;">
      <div style="font-size:13px; color:#713f12; line-height:1.5;">
        The <a href="https://www.congress.gov/crs-product/IN12565" target="_blank" style="color:#854d0e; font-weight:600;">FY2026 Agriculture Appropriations Act</a>
        (signed Nov 2025) redefines hemp to include <b>total THC</b>
        and caps finished products at <b>0.4mg THC per container</b>, effective <b>Nov 12, 2026</b>.
        This effectively bans most hemp-derived THC products currently on the market. The alternative CSRA bill
        proposes a more permissive framework (5mg/serving, 50mg/container). Current market data represents a
        snapshot of a market that may fundamentally change within months.
        <div style="margin-top:10px; padding:10px; background:rgba(255,255,255,0.5); border-radius:6px; font-size:12px; line-height:1.6;">
          <b>If 0.4mg rule takes effect:</b> The hemp-derived THC gummy channel ({fmt_b(tam_hemp_lo)}&ndash;{fmt_b(tam_hemp_hi)}) effectively goes to zero in current form.
          Some revenue migrates to licensed dispensary (states with legal rec), some reformulates to CBD/minor cannabinoids,
          and some shifts to the supplement gummy market. Net TAM could shrink to ~{fmt_b(tam_disp + tam_amz)} (dispensary + Amazon only).<br>
          <b>If CSRA passes instead:</b> 5mg/serving, 50mg/container preserves most current products. TAM largely intact with modest reformulation.
        </div>
      </div>
    </div>
  </details>"""

        market_context_html += f"""

  <div class="card full-width" style="padding:24px;">
    <h3 style="font-size:16px; margin-bottom:4px; color:#2d3748;">Total Addressable Market — Gummies</h3>
    <div style="font-size:12px; color:#718096; margin-bottom:16px;">Top-down model combining BDSA, Headset, and Brightfield industry data</div>

    <div style="display:flex; gap:16px; flex-wrap:wrap; margin-bottom:20px;">
      <div style="flex:1; min-width:180px; background:linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%); border-radius:8px; padding:16px;">
        <div style="font-size:11px; color:#3b82f6; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;">Total Gummy Market</div>
        <div style="font-size:28px; font-weight:800; color:#1e40af; margin:4px 0;">{fmt_b(tam_lo)} – {fmt_b(tam_hi)}</div>
        <div style="font-size:11px; color:#6b7280;">Best est. ~{fmt_b((tam_lo + tam_hi) / 2)} &mdash; Brightfield (low) vs 2&times; Brightfield (high)</div>
      </div>
      <div style="flex:1; min-width:180px; background:linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%); border-radius:8px; padding:16px;">
        <div style="font-size:11px; color:#16a34a; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;">Functional Gummy TAM</div>
        <div style="font-size:28px; font-weight:800; color:#15803d; margin:4px 0;">{fmt_b(func_lo)} – {fmt_b(func_hi)}</div>
        <div style="font-size:11px; color:#6b7280;">Sleep, pain, relaxation, focus, mood &mdash; per-channel functional shares</div>
      </div>
      <div style="flex:1; min-width:180px; background:linear-gradient(135deg, #fefce8 0%, #fef9c3 100%); border-radius:8px; padding:16px;">
        <div style="font-size:11px; color:#a16207; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;">Growth &amp; Position</div>
        <div style="font-size:28px; font-weight:800; color:#854d0e; margin:4px 0;">+7.3%<span style="font-size:14px; font-weight:400;"> YoY</span></div>
        <div style="font-size:11px; color:#6b7280;">Hemp gummies 2024&rarr;2025 (Brightfield)</div>
        <div style="font-size:10px; color:#9ca3af; margin-top:2px;">Mood = ~1.5% of total hemp THC market (Brightfield company rankings)</div>
      </div>
    </div>

    <table style="width:100%; font-size:13px; border-collapse:collapse; margin-bottom:16px;">
      <thead><tr style="border-bottom:2px solid #e2e8f0;">
        <th style="padding:8px 12px; text-align:left; font-weight:600;">Channel</th>
        <th style="padding:8px 12px; text-align:right; font-weight:600;">Gummy TAM</th>
        <th style="padding:8px 12px; text-align:right; font-weight:600;">Func. %</th>
        <th style="padding:8px 12px; text-align:right; font-weight:600;">Func. TAM</th>
        <th style="padding:8px 12px; text-align:left; font-weight:600;">Formula</th>
      </tr></thead>
      <tbody style="color:#374151;">
        <tr><td style="padding:8px 12px;">Licensed dispensary</td>
            <td style="padding:8px 12px; text-align:right; font-weight:600;">{fmt_b(tam_disp)}</td>
            <td style="padding:8px 12px; text-align:right;">{fs_disp:.0%}</td>
            <td style="padding:8px 12px; text-align:right; font-weight:600;">{fmt_b(tam_disp * fs_disp)}</td>
            <td style="padding:8px 12px; font-size:11px;">
              <span style="font-family:monospace; background:#f1f5f9; padding:1px 4px; border-radius:2px;">$4.3B &times; 73% = {fmt_b(tam_disp)}</span>
              <details style="margin-top:4px;"><summary style="cursor:pointer; color:#2563EB; font-size:10px;">Show derivation</summary>
              <div style="margin-top:4px; padding:6px 8px; background:#f8fafc; border-radius:4px; font-size:10px; color:#718096; line-height:1.6;">
                <span style="color:#374151; font-weight:500;">$4.3B</span> = US regulated cannabis edibles sales (2025)<br>
                &nbsp;&nbsp;Source: <a href="https://www.cannabissciencetech.com/view/2025-state-of-the-cannabis-industry-sales-trends-and-forecasts" target="_blank" style="color:#2563EB;">BDSA &ldquo;2025 State of the Cannabis Industry&rdquo;</a><br>
                &nbsp;&nbsp;Context: #3 category after flower ($11.8B) and vapes ($7.7B)<br><br>
                <span style="color:#374151; font-weight:500;">73%</span> = gummies share of edible sales<br>
                &nbsp;&nbsp;Source: <a href="https://www.headset.io/industry-reports/cannabis-edibles-an-analysis-of-category-trends-performance" target="_blank" style="color:#2563EB;">Headset &ldquo;Cannabis Edibles: Category Trends&rdquo; (2022)</a>: 72&ndash;77% range, midpoint used<br>
                &nbsp;&nbsp;Cross-check: <a href="https://bdsa.com/bdsa-webinar-recap-eat-insights-into-the-edibles-market/" target="_blank" style="color:#2563EB;">BDSA &ldquo;EAT: Insights into Edibles&rdquo; (Q1 2025)</a>: candy = 79% of edibles, gummies = 85% of candy &rarr; 67%<br>
                &nbsp;&nbsp;Also: <a href="https://www.flowhub.com/cannabis-industry-statistics" target="_blank" style="color:#2563EB;">Flowhub 2026 Statistics</a>: gummies = 72% of edibles
              </div></details></td></tr>
        <tr><td style="padding:8px 12px;">Hemp-derived THC</td>
            <td style="padding:8px 12px; text-align:right; font-weight:600;">{fmt_b(tam_hemp_lo)} – {fmt_b(tam_hemp_hi)}</td>
            <td style="padding:8px 12px; text-align:right;">{fs_dtc:.0%}</td>
            <td style="padding:8px 12px; text-align:right; font-weight:600;">{fmt_b(tam_hemp_lo * fs_dtc)} – {fmt_b(tam_hemp_hi * fs_dtc)}</td>
            <td style="padding:8px 12px; font-size:11px;">
              <b>Low:</b> <span style="font-family:monospace; background:#f1f5f9; padding:1px 4px; border-radius:2px;">{fmt_b(tam_hemp_lo)}</span> &nbsp;
              <b>High:</b> <span style="font-family:monospace; background:#f1f5f9; padding:1px 4px; border-radius:2px;">2 &times; {fmt_b(tam_hemp_lo)} = {fmt_b(tam_hemp_hi)}</span>
              <details style="margin-top:4px;"><summary style="cursor:pointer; color:#2563EB; font-size:10px;">Show derivation</summary>
              <div style="margin-top:4px; padding:6px 8px; background:#f8fafc; border-radius:4px; font-size:10px; color:#718096; line-height:1.6;">
                <b style="color:#374151;">Low estimate ({fmt_b(tam_hemp_lo)}):</b> Hemp-derived THC gummies market size directly from
                <a href="https://app.brightfieldgroup.com/pages/consumer-hemp-derived-thc" target="_blank" style="color:#2563EB;">Brightfield Group Hemp-Derived THC Report (Nov 2025)</a>.
                Brightfield sizes the total hemp THC market at $5.42B; gummies = ${tam_hemp_lo/1e9:.2f}B of that.<br><br>
                <b style="color:#374151;">High estimate ({fmt_b(tam_hemp_hi)}):</b> 2&times; Brightfield to account for undercounting in
                unregulated channels (smoke shops, gas stations, convenience stores) that Brightfield may underweight.<br><br>
                <b style="color:#92400e;">Why not higher?</b> BDSA sizes the total intoxicating hemp market at $21.8B, which would imply
                ~$3.9B in gummies alone ($21.8B &times; 27% edibles &times; 67% gummies). But this would mean unregulated hemp is
                <b>4&times; the entire regulated dispensary market</b> ($4.3B) &mdash; an extraordinary claim that lacks independent confirmation.
                We cap the high estimate at 2&times; Brightfield as a more defensible upper bound.
              </div></details></td></tr>
        <tr><td style="padding:8px 12px;">Amazon hemp/CBD</td>
            <td style="padding:8px 12px; text-align:right; font-weight:600;">{fmt_b(tam_amz)}</td>
            <td style="padding:8px 12px; text-align:right;">{fs_amz:.0%}</td>
            <td style="padding:8px 12px; text-align:right; font-weight:600;">{fmt_b(tam_amz * fs_amz)}</td>
            <td style="padding:8px 12px; font-size:11px;">
              <span style="font-family:monospace; background:#f1f5f9; padding:1px 4px; border-radius:2px;">({fmt_rev(amz_measured)} + {fmt_rev(amz_extrapolated)}) &times; 12</span>
              <details style="margin-top:4px;"><summary style="cursor:pointer; color:#2563EB; font-size:10px;">Show derivation</summary>
              <div style="margin-top:4px; padding:6px 8px; background:#f8fafc; border-radius:4px; font-size:10px; color:#718096; line-height:1.6;">
                <span style="color:#374151; font-weight:500;">{fmt_rev(amz_measured)}/mo measured:</span> {ms.get('amazon_measured_products', 0):.0f} products with Amazon &ldquo;bought in past month&rdquo; badges &times; listed price<br>
                &nbsp;&nbsp;Data: Playwright scrape of Amazon search results (Apr 2026), {amz_products} total gummy products<br><br>
                <span style="color:#374151; font-weight:500;">{fmt_rev(amz_extrapolated)}/mo extrapolated:</span> {amz_products - ms.get('amazon_measured_products', 0):.0f} products without sold badges &times; 30 est. units/mo &times; median unbadged price<br>
                &nbsp;&nbsp;Assumption: unbadged products sell ~30 units/mo on average (low volume, below Amazon&rsquo;s display threshold)
              </div></details></td></tr>
        <tr style="border-top:2px solid #e2e8f0; font-weight:700;">
            <td style="padding:8px 12px;">Total</td>
            <td style="padding:8px 12px; text-align:right;">{fmt_b(tam_lo)} – {fmt_b(tam_hi)}</td>
            <td style="padding:8px 12px;"></td>
            <td style="padding:8px 12px; text-align:right;">{fmt_b(func_lo)} – {fmt_b(func_hi)}</td>
            <td style="padding:8px 12px;"></td></tr>
      </tbody>
    </table>

    <div style="margin-bottom:16px; padding:16px; background:linear-gradient(135deg, #faf5ff 0%, #f3e8ff 100%); border-radius:8px; border-left:4px solid #7c3aed;">
      <div style="font-weight:700; font-size:13px; color:#5b21b6; margin-bottom:8px;">Broader Context: Supplement Gummies Compete for the Same Consumer</div>
      <div style="font-size:12px; color:#374151; line-height:1.6; margin-bottom:10px;">
        The US supplement gummy market (melatonin, ashwagandha, nootropics) is <b>$4.2B&ndash;$5.5B</b> (2024, GM Insights / Fortune Business Insights),
        growing at <b>10&ndash;13% CAGR</b>. Supplement gummies sell at Walmart, Target, CVS, and Amazon; their market size is based on
        <b>retail scanner data</b> (IRI/SPINS/Nielsen). Our cannabis/hemp TAM above is a <b>modeled estimate</b> combining dispensary POS data
        with Brightfield survey-based sizing &mdash; less precise. The two numbers are not directly comparable, but both markets
        target the same functional needs.
      </div>
      <table style="width:100%; font-size:11px; border-collapse:collapse; margin-bottom:8px;">
        <tr style="border-bottom:1px solid #e2e8f0;">
          <td style="padding:4px 8px; font-weight:600;">Segment</td>
          <td style="padding:4px 8px; font-weight:600; text-align:right;">Supplement Gummies (US)</td>
          <td style="padding:4px 8px; font-weight:600;">Competes with</td>
        </tr>
        <tr style="border-bottom:1px solid #f0f0f0;">
          <td style="padding:4px 8px;">Sleep (melatonin)</td>
          <td style="padding:4px 8px; text-align:right; font-weight:600;">~$950M</td>
          <td style="padding:4px 8px; color:#718096;">THC/CBN sleep gummies (Wyld, Mood, Plus)</td>
        </tr>
        <tr style="border-bottom:1px solid #f0f0f0;">
          <td style="padding:4px 8px;">Stress/calm (ashwagandha)</td>
          <td style="padding:4px 8px; text-align:right; font-weight:600;">$500M&ndash;$2.2B</td>
          <td style="padding:4px 8px; color:#718096;">THC/CBD relaxation gummies</td>
        </tr>
        <tr style="border-bottom:1px solid #f0f0f0;">
          <td style="padding:4px 8px;">Focus/cognitive (nootropics)</td>
          <td style="padding:4px 8px; text-align:right; font-weight:600;">$530M&ndash;$710M</td>
          <td style="padding:4px 8px; color:#718096;">Microdose/focus gummies</td>
        </tr>
        <tr>
          <td style="padding:4px 8px;">Energy (caffeine, B vitamins)</td>
          <td style="padding:4px 8px; text-align:right; font-weight:600;">$1.6B&ndash;$2.5B</td>
          <td style="padding:4px 8px; color:#718096;">THC energy/sativa gummies</td>
        </tr>
      </table>
      <div style="font-size:11px; color:#6b7280;">
        <b>Total functional overlap: ~$3.6B&ndash;$6.7B</b> in supplement gummies that compete for the same consumer needs.
        Combined addressable market (cannabis + supplement functional gummies): <b>~$8B&ndash;$12B</b>.
        <br><span style="font-size:10px; color:#9ca3af;"><b>Caveat:</b> Supplement figures are retail-scanner-measured; cannabis/hemp TAM is modeled. Ranges overlap, not additive.</span>
        <br>Key players: Olly (~$500M revenue, #1 gummy brand), Vitafusion (18&ndash;22% share), ZzzQuil Pure Zzzs (#1 sleep aid).
      </div>
      <details style="margin-top:6px;"><summary style="cursor:pointer; color:#7c3aed; font-size:10px;">Sources</summary>
      <div style="font-size:10px; color:#9ca3af; margin-top:4px; line-height:1.5;">
        <a href="https://www.gminsights.com/industry-analysis/gummy-supplements-market" target="_blank" style="color:#7c3aed;">GM Insights</a> (US gummy supplements $4.2B),
        <a href="https://www.fortunebusinessinsights.com/gummy-supplements-market-109478" target="_blank" style="color:#7c3aed;">Fortune Business Insights</a> ($10.78B by 2032),
        <a href="https://www.gminsights.com/industry-analysis/melatonin-supplements-market" target="_blank" style="color:#7c3aed;">GM Insights Melatonin</a> ($950M US),
        <a href="https://www.grandviewresearch.com/industry-analysis/ashwagandha-supplements-market-report" target="_blank" style="color:#7c3aed;">Grand View Research Ashwagandha</a>,
        <a href="https://www.grandviewresearch.com/industry-analysis/us-brain-health-supplements-market-report" target="_blank" style="color:#7c3aed;">Grand View Research Brain Health</a> ($3.56B all forms).
        Gummy format shares estimated where not directly reported.
      </div></details>
    </div>

    <details style="margin-bottom:12px;"><summary style="cursor:pointer; font-weight:700; font-size:12px; color:#2d3748; padding:8px 0;">Functional Share by Channel &mdash; click to expand</summary>
    <div style="font-size:11px; line-height:1.7; padding:12px; background:#f9fafb; border-radius:6px; margin-top:4px;">
      <table style="width:100%; font-size:11px; border-collapse:collapse; margin-bottom:8px;">
        <tr style="border-bottom:1px solid #e2e8f0;">
          <td style="padding:4px 8px; font-weight:600; width:20%;">Channel</td>
          <td style="padding:4px 8px; font-weight:600; width:10%; text-align:right;">Func. %</td>
          <td style="padding:4px 8px; font-weight:600;">How derived</td>
        </tr>
        <tr style="border-bottom:1px solid #f0f0f0;">
          <td style="padding:4px 8px;">Dispensary</td>
          <td style="padding:4px 8px; text-align:right;">{fs_disp:.0%}</td>
          <td style="padding:4px 8px; color:#718096;">Estimated from
            <a href="https://www.headset.io/industry-reports/cannabis-edibles-an-analysis-of-category-trends-performance" target="_blank" style="color:#2563EB;">Headset</a> SKU-level data:
            sleep/CBN gummies (Wyld Elderberry CBN, Plus Deep Sleep, Dialed In Sleep) are consistently top sellers
            (<a href="https://mjbizdaily.com/gummies-dominate-cannabis-infused-edibles-sector-with-two-major-players-vying-for-lead/" target="_blank" style="color:#2563EB;">MJBizDaily</a>),
            but most dispensary gummies are sold for recreational use. 30% is a conservative estimate.</td>
        </tr>
        <tr style="border-bottom:1px solid #f0f0f0;">
          <td style="padding:4px 8px;">Hemp DTC</td>
          <td style="padding:4px 8px; text-align:right;">{fs_dtc:.0%}</td>
          <td style="padding:4px 8px; color:#718096;">Computed from our {dtc_prods_n:,}-product DTC dataset (scraped from {ms.get('dtc_brands_count', 0)} brand Shopify stores, Apr 2026).
            {fs_dtc:.0%} of gummy products include a functional claim (sleep, relaxation, pain, focus, mood, etc.) in their title or tags.
            Classified using keyword matching against 11 effect categories.</td>
        </tr>
        <tr>
          <td style="padding:4px 8px;">Amazon</td>
          <td style="padding:4px 8px; text-align:right;">{fs_amz:.0%}</td>
          <td style="padding:4px 8px; color:#718096;">Computed from our {amz_products}-product Amazon dataset (PLP scrape, Apr 2026).
            Revenue-weighted: {fs_amz:.0%} of measured Amazon gummy revenue ({fmt_rev(amz_measured)}/mo) comes from products with functional claims.
            Higher than product-count share because functional gummies have higher sell-through.</td>
        </tr>
      </table>
    </div></details>

    <details><summary style="cursor:pointer; font-weight:700; font-size:12px; color:#2d3748; padding:8px 0;">Not Included in TAM &mdash; click to expand</summary>
    <div style="padding:8px 12px; margin-top:4px; background:#fef2f2; border-radius:6px; color:#92400e; font-size:10px; line-height:1.6;">
      &bull; <b>Retail pharmacy/grocery CBD gummies</b> (Walgreens, CVS, grocery &mdash; tracked by
      <a href="https://www.spins.com/" target="_blank" style="color:#2563EB;">SPINS</a>, est. $500M&ndash;$1B)<br>
      &bull; <b>Gas station/convenience store hemp THC</b> (large untracked channel, growing rapidly)<br>
      &bull; <b>International markets</b> (Canada, EU emerging &mdash; all numbers above are US only)<br>
      &bull; <b>Smoke shops/head shops</b> (significant hemp THC distribution, no public data)
    </div></details>
  </div>"""

        # ── What We Measure card ──────────────────────────────────────
        market_context_html += f"""

  <div class="card" style="padding:24px;">
    <h3 style="font-size:16px; margin-bottom:4px; color:#2d3748;">What We Measure</h3>
    <div style="font-size:11px; color:#718096; margin-bottom:12px;">Bottom-up from tracked products across {marketplace_count} channels</div>
    <div style="font-size:28px; font-weight:800; color:#2563EB; margin-bottom:4px;">{fmt_rev(total_tracked_lo)} – {fmt_rev(total_tracked_hi)}<span style="font-size:14px; font-weight:400; color:#718096;">/mo est.</span></div>
    <div style="font-size:13px; color:#718096; margin-bottom:16px;">Across {amz_products + dtc_products:,} products from {total_brands} brands</div>
    <table style="width:100%; font-size:12px; border-collapse:collapse; table-layout:fixed;">
      <colgroup><col style="width:30%"><col style="width:30%"><col style="width:40%"></colgroup>
      <thead><tr style="border-bottom:2px solid #e2e8f0;">
        <th style="padding:6px 8px; text-align:left; font-weight:600;">Channel</th>
        <th style="padding:6px 8px; text-align:right; font-weight:600;">Revenue</th>
        <th style="padding:6px 8px; text-align:left; font-weight:600;">Confidence</th>
      </tr></thead>
      <tbody style="color:#374151;">
        <tr><td style="padding:6px 8px;">Amazon (measured)</td>
            <td style="padding:6px 8px; text-align:right; font-weight:600;">{fmt_rev(amz_measured)}/mo</td>
            <td style="padding:6px 8px;"><span style="background:#dcfce7; color:#166534; padding:2px 6px; border-radius:3px; font-size:11px; font-weight:600;">Measured</span><br><span style="font-size:10px; color:#718096;">{amz_pct:.0%} have badges</span></td></tr>
        <tr><td style="padding:6px 8px;">Amazon (long-tail est.)</td>
            <td style="padding:6px 8px; text-align:right; font-weight:600;">{fmt_rev(amz_extrapolated)}/mo</td>
            <td style="padding:6px 8px;"><span style="background:#fef9c3; color:#854d0e; padding:2px 6px; border-radius:3px; font-size:11px; font-weight:600;">Estimated</span><br><span style="font-size:10px; color:#718096;">30 units/mo assumed</span></td></tr>"""

        if dtc_est > 0:
            market_context_html += f"""
        <tr><td style="padding:6px 8px;">DTC (traffic model)</td>
            <td style="padding:6px 8px; text-align:right; font-weight:600;">{fmt_rev(dtc_lo)} – {fmt_rev(dtc_hi)}/mo</td>
            <td style="padding:6px 8px;"><span style="background:#fef2f2; color:#991b1b; padding:2px 6px; border-radius:3px; font-size:11px; font-weight:600;">Modeled</span><br><span style="font-size:10px; color:#718096;">{ms.get('dtc_brands_with_traffic', 0)} brands, 5x range</span></td></tr>"""
        elif dtc_brands_n > 0:
            market_context_html += f"""
        <tr><td style="padding:6px 8px;">DTC (not estimated)</td>
            <td style="padding:6px 8px; text-align:right; font-weight:600; color:#718096;">&mdash;</td>
            <td style="padding:6px 8px;"><span style="background:#f1f5f9; color:#475569; padding:2px 6px; border-radius:3px; font-size:11px; font-weight:600;">No data</span><br><span style="font-size:10px; color:#718096;">{dtc_brands_n} brands tracked</span></td></tr>"""

        if disp_est > 0:
            market_context_html += f"""
        <tr><td style="padding:6px 8px;">Dispensary (rough est.)</td>
            <td style="padding:6px 8px; text-align:right; font-weight:600;">{fmt_rev(disp_est)}/mo</td>
            <td style="padding:6px 8px;"><span style="background:#fef2f2; color:#991b1b; padding:2px 6px; border-radius:3px; font-size:11px; font-weight:600;">Rough</span><br><span style="font-size:10px; color:#718096;">{sum(1 for p in products if p['source'] == 'Dispensary')} products &times; 100 units</span></td></tr>"""

        market_context_html += f"""
      <tr style="border-top:2px solid #e2e8f0; font-weight:700;">
        <td style="padding:6px 8px;">Total</td>
        <td style="padding:6px 8px; text-align:right;">{fmt_rev(total_tracked_lo)} – {fmt_rev(total_tracked_hi)}/mo</td>
        <td style="padding:6px 8px;"></td>
      </tr></tbody>
    </table>
    <div style="margin-top:12px; padding:10px; background:#f0fdf4; border-radius:6px; font-size:12px; color:#166534;">
      Our tracked <b>{fmt_rev(total_tracked_lo * 12)}–{fmt_rev(total_tracked_hi * 12)}/yr</b> represents <b>{tam_coverage_lo:.1f}–{tam_coverage_hi:.1f}%</b> of the
      {fmt_b(tam_lo)}–{fmt_b(tam_hi)} total gummy TAM. Amazon and DTC are two channels in one country;
      dispensary, convenience/gas station, and international channels account for the vast majority.
    </div>
  </div>

  <div class="card" style="padding:24px;">
    <h3 style="font-size:16px; margin-bottom:16px; color:#2d3748;">TAM Sources &amp; How They Add Up</h3>
    <div style="font-size:13px; line-height:1.6; color:#374151;">"""

        # Use Brightfield data if available
        bf = brightfield or {}
        gummies_2025 = bf.get("gummies_2025")
        gummies_2024 = bf.get("gummies_2024")
        gummies_2030 = bf.get("gummies_2030")
        total_mkt_2025 = bf.get("total_market_2025")

        # Dispensary source (always BDSA)
        bdsa_link = "https://www.cannabissciencetech.com/view/2025-state-of-the-cannabis-industry-sales-trends-and-forecasts"
        headset_link = "https://www.headset.io/industry-reports/cannabis-edibles-an-analysis-of-category-trends-performance"
        market_context_html += f"""
      <div style="font-weight:700; font-size:11px; color:#6b7280; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:4px;">1. Licensed Dispensary &mdash; {fmt_b(tam_disp)}</div>
      <div style="display:flex; justify-content:space-between; padding:4px 0; border-bottom:1px solid #f0f0f0; font-size:12px;">
        <span><a href="{bdsa_link}" target="_blank" style="color:#2563EB;">US regulated edibles (BDSA 2025)</a></span><span style="font-weight:600;">$4.3B</span>
      </div>
      <div style="display:flex; justify-content:space-between; padding:4px 0; border-bottom:1px solid #f0f0f0; font-size:12px;">
        <span><a href="{headset_link}" target="_blank" style="color:#2563EB;">Gummies share of edibles (Headset)</a></span><span style="font-weight:600;">73%</span>
      </div>"""

        if gummies_2025:
            gummies_yoy = ((gummies_2025 / gummies_2024 - 1) * 100) if gummies_2024 else 0
            bf_link = "https://app.brightfieldgroup.com/pages/consumer-hemp-derived-thc"
            market_context_html += f"""
      <div style="font-weight:700; font-size:11px; color:#6b7280; text-transform:uppercase; letter-spacing:0.5px; margin-top:12px; margin-bottom:4px;">2. Hemp-Derived THC &mdash; {fmt_b(tam_hemp_lo)}&ndash;{fmt_b(tam_hemp_hi)}</div>
      <div style="display:flex; justify-content:space-between; padding:4px 0; border-bottom:1px solid #f0f0f0; font-size:12px;">
        <span><a href="{bf_link}" target="_blank" style="color:#2563EB;">Hemp THC gummies (Brightfield 2025)</a></span><span style="font-weight:700; color:#2563EB;">{fmt_b(tam_hemp_lo)}</span>
      </div>
      <div style="display:flex; justify-content:space-between; padding:4px 0; border-bottom:1px solid #f0f0f0; font-size:12px;">
        <span>High estimate (2&times; Brightfield)</span><span style="font-weight:600;">{fmt_b(tam_hemp_hi)}</span>
      </div>
      <div style="display:flex; justify-content:space-between; padding:4px 0; border-bottom:1px solid #f0f0f0; font-size:12px;">
        <span><a href="{bf_link}" target="_blank" style="color:#2563EB;">YoY growth (2024&rarr;2025)</a></span><span style="font-weight:600;">{gummies_yoy:+.1f}%</span>
      </div>
      <div style="display:flex; justify-content:space-between; padding:4px 0; border-bottom:1px solid #f0f0f0; font-size:12px;">
        <span><a href="{bf_link}" target="_blank" style="color:#2563EB;">Projected 2030</a></span><span style="font-weight:600;">${gummies_2030/1e9:.1f}B</span>
      </div>"""
        else:
            market_context_html += f"""
      <div style="font-weight:700; font-size:11px; color:#6b7280; text-transform:uppercase; letter-spacing:0.5px; margin-top:12px; margin-bottom:4px;">2. Hemp-Derived THC</div>
      <div style="display:flex; justify-content:space-between; padding:4px 0; border-bottom:1px solid #f0f0f0; font-size:12px;">
        <span><a href="https://bdsa.com/bdsa-the-rise-of-intoxicating-hemp-products-and-their-impact-on-the-cannabis-market/" target="_blank" style="color:#2563EB;">Hemp THC edibles (BDSA 2025)</a></span><span style="font-weight:600;">~$5.9B</span>
      </div>"""

        market_context_html += f"""
      <div style="font-weight:700; font-size:11px; color:#6b7280; text-transform:uppercase; letter-spacing:0.5px; margin-top:12px; margin-bottom:4px;">3. Amazon &mdash; {fmt_b(tam_amz)}</div>
      <div style="display:flex; justify-content:space-between; padding:4px 0; border-bottom:1px solid #f0f0f0; font-size:12px;">
        <span>Measured + extrapolated, annualized</span><span style="font-weight:600;">{fmt_b(tam_amz)}</span>
      </div>
      <div style="margin-top:12px; padding:8px; background:#f0f4ff; border-radius:6px; font-size:12px; font-weight:600; color:#1e3a5f; display:flex; justify-content:space-between;">
        <span>Total TAM = (1) + (2) + (3)</span><span>{fmt_b(tam_lo)} &ndash; {fmt_b(tam_hi)}</span>
      </div>
      <div style="margin-top:8px; font-size:10px; color:#9ca3af;">
        Sources: <a href="{bdsa_link}" target="_blank" style="color:#718096;">BDSA</a>,
        <a href="{headset_link}" target="_blank" style="color:#718096;">Headset</a>,
        {"<a href='" + bf_link + "' target='_blank' style='color:#718096;'>Brightfield Group</a>." if gummies_2025 else
         '<a href="https://www.grandviewresearch.com/industry-analysis/cbd-gummies-market" target="_blank" style="color:#718096;">Grand View Research</a>.'}
        Industry estimates are approximate and vary by methodology.
      </div>
    </div>
  </div>"""

    # Product data for table
    table_data = []
    for p in products:
        table_data.append({
            "source": p["source"],
            "id": p.get("id", ""),
            "brand": p.get("brand") or "Unknown",
            "productName": (p.get("productName") or "")[:80],
            "cannabinoids": ", ".join(p.get("cannabinoids", [])),
            "formFactor": p.get("formFactor") or "",
            "effects": ", ".join(p.get("effects", [])),
            "price": p.get("price"),
            "rating": p.get("rating"),
            "reviewCount": p.get("reviewCount"),
            "soldPastMonth": p.get("soldPastMonth"),
            "url": p.get("url") or "",
        })

    source_options = "".join(f'<option value="{s}">{s}</option>'
                             for s in sorted(set(p["source"] for p in products)))
    ff_options = "".join(f'<option value="{ff}">{ff}</option>'
                         for ff in sorted(set(p.get("formFactor") for p in products if p.get("formFactor"))))
    all_cbs = set()
    for p in products:
        for cb in p.get("cannabinoids", []):
            all_cbs.add(cb)
    cb_options = "".join(f'<option value="{cb}">{cb}</option>' for cb in sorted(all_cbs))
    all_effects = set()
    for p in products:
        for eff in p.get("effects", []):
            all_effects.add(eff)
    eff_options = "".join(f'<option value="{e}">{e}</option>' for e in sorted(all_effects))

    table_json = json.dumps(table_data, default=str)

    html = f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Functional Gummies Market Dashboard</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #f7fafc; color: #1a202c; }}
  .header {{ background: linear-gradient(135deg, #1a365d 0%, #2d3748 100%); color: white; padding: 32px 40px; }}
  .header h1 {{ font-size: 28px; margin-bottom: 8px; }}
  .header p {{ opacity: 0.8; font-size: 15px; }}
  .stats {{ display: flex; gap: 24px; margin-top: 20px; flex-wrap: wrap; }}
  .stat {{ background: rgba(255,255,255,0.1); border-radius: 8px; padding: 12px 20px; min-width: 120px; }}
  .stat .num {{ font-size: 22px; font-weight: 700; }}
  .stat .label {{ font-size: 12px; opacity: 0.7; margin-top: 2px; }}
  .grid {{ max-width: 1400px; margin: 24px auto; padding: 0 24px; display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
  .card {{ background: white; border-radius: 12px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); overflow: hidden; }}
  .card.full-width {{ grid-column: 1 / -1; }}
  .section-header {{ grid-column: 1 / -1; margin-top: 16px; }}
  .section-header h2 {{ font-size: 20px; color: #2d3748; }}
  .section-header p {{ font-size: 13px; color: #718096; margin-top: 4px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
  th {{ background: #f7fafc; padding: 8px 6px; text-align: left; font-weight: 600; border-bottom: 2px solid #e2e8f0; cursor: pointer; }}
  td {{ padding: 6px; border-bottom: 1px solid #f0f0f0; }}
  tr:hover {{ background: #f7fafc; }}
  a {{ color: #2563EB; text-decoration: none; }}
</style>
</head><body>
<div class="header">
  <h1>Functional Gummies Market Dashboard</h1>
  <p>Hemp-derived THC & CBD functional gummy landscape — sleep, pain, relaxation, focus, intimacy & more</p>
  <div class="stats">
    <div class="stat"><div class="num">{total_products:,}</div><div class="label">Products</div></div>
    <div class="stat"><div class="num">{functional_count:,}</div><div class="label">Functional ({functional_pct:.0%})</div></div>
    <div class="stat"><div class="num">{total_brands:,}</div><div class="label">Brands</div></div>
    <div class="stat"><div class="num">{marketplace_count}</div><div class="label">Channels</div></div>
    <div class="stat"><div class="num">{avg_price}</div><div class="label">Avg Price</div></div>
  </div>
</div>
<div class="grid">
  {market_context_html}
  <div class="card full-width">{charts.get('bf_companies', '')}</div>
  {charts.get('bf_section', '')}
  <div class="section-header"><h2>Functional Categories</h2><p>Products classified by use case — sleep, pain, relaxation, focus, intimacy, cognitive & more</p></div>
  <div class="card full-width">{charts['effect_pop']}</div>

  <div class="section-header"><h2>Market Overview</h2><p>Product landscape across all channels</p></div>
  <div class="card full-width">{charts['brand_map']}</div>
  <div class="card">{charts['cannabinoid_pop']}</div>

  <div class="section-header"><h2>Demand &amp; Revenue</h2><p>Estimated revenue across Amazon (measured) and DTC (modeled from traffic)</p></div>
  <div class="card full-width">{charts['market_map']}</div>
  <div class="card full-width">{charts['dtc_revenue']}</div>
  <div class="card full-width" style="padding:20px;">{charts['dtc_decomposition']}</div>

  <div class="section-header"><h2>Product Explorer</h2><p>Filter and search all {total_products:,} products</p></div>
  <div class="card full-width" style="padding: 20px;">
    <div style="display:flex; flex-wrap:wrap; gap:12px; margin-bottom:16px; align-items:end;">
      <div>
        <label style="font-size:12px; font-weight:600; color:#718096; display:block; margin-bottom:4px;">Search</label>
        <input type="text" id="dt-search" placeholder="Product name, brand..."
               style="padding:8px 12px; border:1px solid #e2e8f0; border-radius:6px; width:260px; font-size:13px;">
      </div>
      <div>
        <label style="font-size:12px; font-weight:600; color:#718096; display:block; margin-bottom:4px;">Channel</label>
        <select id="dt-source" style="padding:8px 12px; border:1px solid #e2e8f0; border-radius:6px; font-size:13px;">
          <option value="">All</option>{source_options}
        </select>
      </div>
      <div>
        <label style="font-size:12px; font-weight:600; color:#718096; display:block; margin-bottom:4px;">Cannabinoid</label>
        <select id="dt-cb" style="padding:8px 12px; border:1px solid #e2e8f0; border-radius:6px; font-size:13px;">
          <option value="">All</option>{cb_options}
        </select>
      </div>
      <div>
        <label style="font-size:12px; font-weight:600; color:#718096; display:block; margin-bottom:4px;">Function</label>
        <select id="dt-eff" style="padding:8px 12px; border:1px solid #e2e8f0; border-radius:6px; font-size:13px;">
          <option value="">All</option><option value="_functional">Any Functional</option>{eff_options}
        </select>
      </div>
    </div>
    <div style="overflow-x:auto;"><table id="product-table">
      <thead><tr>
        <th onclick="sortTable(0)">Source</th>
        <th onclick="sortTable(1)">Brand</th>
        <th onclick="sortTable(2)">Product</th>
        <th onclick="sortTable(3)">Cannabinoids</th>
        <th onclick="sortTable(4)">Effects</th>
        <th onclick="sortTable(5)">Price</th>
        <th onclick="sortTable(6)">Rating</th>
        <th onclick="sortTable(7)">Reviews</th>
        <th onclick="sortTable(8)">Sold/Mo</th>
      </tr></thead>
      <tbody id="dt-body"></tbody>
    </table></div>
    <div id="dt-count" style="margin-top:10px; font-size:13px; color:#718096;"></div>
  </div>
</div>

<script>
const allProducts = {table_json};
const srcColors = {{"Amazon":"#FF9900","DTC":"#9B59B6","Faire":"#5B63FE","Dispensary":"#2ECC71"}};

function renderTable() {{
  const search = document.getElementById('dt-search').value.toLowerCase();
  const src = document.getElementById('dt-source').value;
  const cb = document.getElementById('dt-cb').value;
  const eff = document.getElementById('dt-eff').value;
  let filtered = allProducts.filter(p => {{
    if (src && p.source !== src) return false;
    if (cb && !(p.cannabinoids || '').includes(cb)) return false;
    if (eff === '_functional' && !(p.effects || '').length) return false;
    if (eff && eff !== '_functional' && !(p.effects || '').includes(eff)) return false;
    if (search) {{
      const hay = (p.productName + ' ' + p.brand + ' ' + p.cannabinoids + ' ' + p.effects).toLowerCase();
      if (!hay.includes(search)) return false;
    }}
    return true;
  }});
  filtered.sort((a,b) => (b.soldPastMonth||0) - (a.soldPastMonth||0) || (b.reviewCount||0) - (a.reviewCount||0));
  const show = filtered.slice(0, 200);
  const tbody = document.getElementById('dt-body');
  tbody.innerHTML = show.map(p => `<tr>
    <td style="padding:6px;"><span style="background:${{srcColors[p.source]||'#999'}};color:white;padding:2px 8px;border-radius:4px;font-size:11px;">${{p.source}}</span></td>
    <td style="padding:6px;">${{p.brand}}</td>
    <td style="padding:6px;">${{p.url ? '<a href="'+p.url+'" target="_blank">'+p.productName+'</a>' : p.productName}}</td>
    <td style="padding:6px; font-size:11px;">${{p.cannabinoids||'—'}}</td>
    <td style="padding:6px; font-size:11px;">${{p.effects||'—'}}</td>
    <td style="padding:6px; white-space:nowrap;">${{p.price ? '$'+Number(p.price).toFixed(2) : '—'}}</td>
    <td style="padding:6px;">${{p.rating ? Number(p.rating).toFixed(1)+'★' : '—'}}</td>
    <td style="padding:6px;">${{p.reviewCount ? p.reviewCount.toLocaleString() : '—'}}</td>
    <td style="padding:6px;">${{p.soldPastMonth ? p.soldPastMonth.toLocaleString() : '—'}}</td>
  </tr>`).join('');
  document.getElementById('dt-count').textContent = `Showing ${{show.length}} of ${{filtered.length}} products`;
}}
document.getElementById('dt-search').addEventListener('input', renderTable);
document.getElementById('dt-source').addEventListener('change', renderTable);
document.getElementById('dt-cb').addEventListener('change', renderTable);
document.getElementById('dt-eff').addEventListener('change', renderTable);
renderTable();

let sortCol = -1, sortAsc = true;
function sortTable(col) {{
  if (sortCol === col) sortAsc = !sortAsc;
  else {{ sortCol = col; sortAsc = true; }}
  const keys = ['source','brand','productName','cannabinoids','effects','price','rating','reviewCount','soldPastMonth'];
  const key = keys[col];
  allProducts.sort((a,b) => {{
    let va = a[key], vb = b[key];
    if (typeof va === 'number' || typeof vb === 'number') {{
      va = va || 0; vb = vb || 0;
      return sortAsc ? va - vb : vb - va;
    }}
    va = (va||'').toString(); vb = (vb||'').toString();
    return sortAsc ? va.localeCompare(vb) : vb.localeCompare(va);
  }});
  renderTable();
}}
</script>
</body></html>"""
    return html


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("Loading data…")
    products = load_products()

    print("Loading Brightfield data…")
    brightfield = load_brightfield_data()

    print("Generating charts…")
    charts = {
        "cannabinoid_pop":    fig_to_html(chart_cannabinoid_popularity(products)),
        "effect_pop":         fig_to_html(chart_effect_popularity(products)),
        "functional_share":   fig_to_html(chart_functional_share(products)),
        "functional_revenue": fig_to_html(chart_functional_revenue(products)),
        "functional_map":     fig_to_html(chart_functional_market_map(products)),
        "price_dist":         fig_to_html(chart_price_distribution(products)),
        "top_brands":         fig_to_html(chart_top_brands(products)),
        "brand_map":          fig_to_html(chart_brand_map(products)),
        "market_map":         fig_to_html(chart_market_map(products)),
        "top_revenue":        fig_to_html(chart_top_revenue(products)),
        "dtc_revenue":        "",  # placeholder, filled after market_size computed
    }

    # Brightfield industry charts
    bf_demand_supply = fig_to_html(chart_demand_vs_supply(products))
    if brightfield:
        bf_growth = fig_to_html(chart_brightfield_market_growth(brightfield))
        bf_cannabinoids = fig_to_html(chart_brightfield_cannabinoid_trends(brightfield))
        bf_companies = fig_to_html(chart_brightfield_company_share(brightfield))
        charts["bf_companies"] = bf_companies
        charts["bf_section"] = f"""
  <div class="section-header"><h2>Industry Data — Brightfield Group</h2><p>Hemp-derived THC market sizing, cannabinoid trends & competitive landscape (2020–2030)</p></div>
  <div class="card full-width">{bf_demand_supply}</div>
  <div class="card full-width">{bf_growth}</div>
  <div class="card full-width">{bf_cannabinoids}</div>"""
    else:
        charts["bf_section"] = f"""
  <div class="section-header"><h2>Consumer Demand — Brightfield Group</h2><p>What consumers want vs what's on shelves</p></div>
  <div class="card full-width">{bf_demand_supply}</div>"""

    print("Computing market size…")
    traffic = load_traffic_data()
    # Pass Brightfield gummies number to TAM model
    if brightfield and brightfield.get("gummies_2025"):
        compute_market_size._brightfield_gummies = brightfield["gummies_2025"]
    market_size = compute_market_size(products, traffic)
    print(f"  Amazon measured: ${market_size['amazon_measured']:,.0f}/mo ({market_size['amazon_measured_pct']:.0%} coverage)")
    print(f"  Amazon extrapolated: ${market_size['amazon_extrapolated']:,.0f}/mo")
    if market_size['dtc_estimated']:
        print(f"  DTC estimated: ${market_size['dtc_estimated']:,.0f}/mo ({market_size['dtc_brands_with_traffic']} brands)")
    if market_size['dispensary_estimated']:
        print(f"  Dispensary estimated: ${market_size['dispensary_estimated']:,.0f}/mo")
    print(f"  Total tracked: ${market_size['total_tracked']:,.0f}/mo")
    print(f"  TAM (gummies): ${market_size['tam_total_low']/1e9:.1f}B – ${market_size['tam_total_high']/1e9:.1f}B")
    print(f"  Functional TAM: ${market_size['func_tam_low']/1e9:.1f}B – ${market_size['func_tam_high']/1e9:.1f}B")

    # DTC revenue chart + decomposition table (needs market_size)
    charts["dtc_revenue"] = fig_to_html(chart_dtc_revenue(market_size))
    charts["dtc_decomposition"] = build_dtc_decomposition_html(market_size, brightfield=brightfield)

    print("Building HTML…")
    html = build_html(products, charts, market_size=market_size, brightfield=brightfield)

    output = "thc_gummies_dashboard.html"
    with open(output, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"\n✅ Dashboard saved → {output}")
    print(f"   Open in browser: open {output}")


if __name__ == "__main__":
    main()
